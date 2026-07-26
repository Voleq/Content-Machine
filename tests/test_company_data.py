"""v3 company-data reader (§3): sheets read by NAME (Snapshot · History ·
Dashboard · Valuation · Peers), History period columns read dynamically
from the header row, CSV = snapshot-only, generic screenshot label."""

import pytest
from PIL import Image

from pipeline.company_data import (
    CompanyDataError,
    FILING_LABEL,
    find_export,
    list_screenshots,
    load_company_data,
    prepare_screenshot,
)
from pipeline.models import HISTORY_FIELDS


def test_load_snapshot_and_history(workspace):
    data = load_company_data(workspace)
    assert data.get("ticker") == "EXMPL"
    assert data.get("ps_ttm") == 62.0
    assert data.get("as_of_date") == "2026-07-01"
    assert data.blocking_missing == []

    assert data.has_history
    # periods are read dynamically from the header row (FY-4..FY-0, LTM)
    assert data.history_years == ["FY-4", "FY-3", "FY-2", "FY-1", "FY-0", "LTM"]
    assert data.history_row("revenue") == [400e6, 452e6, 471e6, 491e6, 496e6, 496e6]
    assert data.history_row("net_income")[-1] == -89e6
    assert set(data.history) <= set(HISTORY_FIELDS)


def test_dashboard_valuation_peers_surfaced(workspace):
    data = load_company_data(workspace)
    # the Dashboard summary — exactly what the numbers sheet reads
    assert data.dashboard_get("Net margin (LTM)") == -18.0
    assert data.dashboard_get("Profitable? (LTM NI>0)") == "no"
    assert data.dashboard_get("Diluting? (share CAGR>1%)") == "yes"
    # long-form context: bear/base/bull + peers
    assert data.has_valuation and len(data.valuation["scenarios"]) == 3
    assert data.valuation["current_price"] == 84.2
    assert data.has_peers and data.peers[0]["name"] == "Freightwave Systems Inc"


def test_news_parsing_skips_blanks_and_coerces_dates(workspace):
    data = load_company_data(workspace)
    # the blank spill row AND the NA-headline row are skipped
    assert len(data.news) == 4
    first = data.news[0]
    assert set(first) == {"date", "headline", "source", "url"}
    # real dates coerce to ISO day strings
    assert first["date"] == "2026-07-14"
    assert first["headline"].startswith("Example Corp posts record")
    assert first["url"] == "https://example.com/news/record-revenue"
    # every surviving row has a non-empty headline
    assert all(n["headline"] for n in data.news)
    # a news outlet as Source is fine; a data-terminal brand never reaches a field
    assert [n["source"] for n in data.news] == [
        "Reuters", "Bloomberg", "Associated Press", "CNBC",
    ]
    blob = " ".join(str(v) for n in data.news for v in n.values()).lower()
    for brand in ("refinitiv", "lseg", "capital iq"):
        assert brand not in blob


def test_peer_percentiles_parsing_and_block_separation(workspace):
    data = load_company_data(workspace)
    pcts = data.peer_percentiles
    assert len(pcts) == 8
    by_metric = {p["metric"]: p for p in pcts}
    # values + a 0-1 percentile fraction + direction + read
    pe = by_metric["P/E (TTM)"]
    assert pe["subject"] == 34.0 and pe["median"] == 37.5
    assert pe["percentile"] == 0.42 and pe["direction"] == "worse"
    assert pe["read"] == "mid-pack"
    rg = by_metric["Rev growth % (3Y CAGR)"]
    assert rg["direction"] == "better" and rg["read"] == "strong vs peers"
    assert rg["percentile"] == 0.80
    assert by_metric["P/S (TTM)"]["read"] == "expensive vs peers"
    assert all(0.0 <= p["percentile"] <= 1.0 for p in pcts)
    assert {p["direction"] for p in pcts} == {"better", "worse"}
    # the two blocks are DISTINCT: the auto peer table did not consume the
    # percentile rows, and the percentile reader did not grab peer names
    assert [p["name"] for p in data.peers] == [
        "Freightwave Systems Inc", "DepotOps Corp",
        "RouteLogic Holdings", "CargoCloud Inc",
    ]
    assert {p["name"] for p in data.peers}.isdisjoint(by_metric)


def test_peer_columns_are_mapped_by_header_not_position(workspace):
    """Pin the peer mapping to the sheet the template actually ships:

        B Price · C Market Cap · D P/E · E EV/EBITDA · F P/S · G Gross Mgn % ·
        H Net Mgn % · I FCF Yield % · J NetDebt/EBITDA · K Rev LTM (now) ·
        L Rev LTM (-3Y) · M Rev Growth % (3Y CAGR)

    The table grew the two revenue feeds and the computed CAGR, pushing
    everything from the margins one column right. Reading fixed offsets B–K
    mislabelled five fields at once — gross margin arriving as `rev_growth`,
    net margin as `gross_margin`, FCF yield as `net_margin`, net debt/EBITDA
    as `fcf_yield`, and a raw revenue figure as `net_debt_ebitda` — each one
    plausible enough on its own to reach a script.
    """
    data = load_company_data(workspace)
    top = data.peers[0]
    assert set(top) == {
        "name", "price", "market_cap", "pe", "ev_ebitda", "ps", "gross_margin",
        "net_margin", "fcf_yield", "net_debt_ebitda", "rev_growth",
    }
    assert top["name"] == "Freightwave Systems Inc"
    assert top["price"] == 142.0
    assert top["market_cap"] == 12_400_000_000
    assert top["pe"] == 34.0
    assert top["ev_ebitda"] == 22.0
    assert top["ps"] == 9.1
    assert top["gross_margin"] == 71.0
    assert top["net_margin"] == 6.0
    assert top["fcf_yield"] == 2.1
    assert top["net_debt_ebitda"] == 0.4
    # rev_growth is the COMPUTED 3Y CAGR (col M) — ((K/L)^(1/3)-1)*100
    assert top["rev_growth"] == pytest.approx(18.0, abs=1e-3)

    # the exact off-by-one the fixed offsets produced, field by field
    assert top["rev_growth"] != 71.0, "gross margin read as revenue growth"
    assert top["gross_margin"] != 6.0, "net margin read as gross margin"
    assert top["net_margin"] != 2.1, "FCF yield read as net margin"
    assert top["fcf_yield"] != 0.4, "net debt/EBITDA read as FCF yield"
    # …and the raw revenue feeds (cols K/L) reach no field at all
    feeds = {2037.36, 1240.0, 1269.13, 980.0, 3813.28, 2100.0, 726.52, 610.0}
    read = {v for p in data.peers for k, v in p.items() if k != "name"}
    assert not (feeds & read), "a raw `Rev LTM` figure landed in a metric field"

    # NA cells still coerce to None, and negative margins survive
    depot = data.peers[1]
    assert depot["name"] == "DepotOps Corp"
    assert depot["pe"] is None                # '#N/A' in the fixture
    assert depot["ev_ebitda"] == 41.0
    assert depot["gross_margin"] == 62.0 and depot["net_margin"] == -4.0
    assert depot["net_debt_ebitda"] == 1.8
    assert depot["rev_growth"] == pytest.approx(9.0, abs=1e-3)
    assert data.peers[3]["fcf_yield"] is None  # '#N/A' — not the neighbour's


def test_peer_columns_are_found_however_the_template_orders_them():
    """Column ORDER is not the contract, the header text is.

    The pre-CAGR revision carried a raw `Rev Growth %` sixth, straight after
    `P/S`; the current one computes a 3Y CAGR at the far right, past two
    revenue feeds. Both must read to the same fields — as must a sheet whose
    columns simply sit in another order, or whose margins are spelled out.
    """
    from openpyxl import Workbook

    from pipeline.company_data import _read_peers

    expected = {
        "name": "Freightwave Systems Inc", "price": 142.0,
        "market_cap": 12_400_000_000, "pe": 34.0, "ev_ebitda": 22.0, "ps": 9.1,
        "gross_margin": 71.0, "net_margin": 6.0, "fcf_yield": 2.1,
        "net_debt_ebitda": 0.4, "rev_growth": 18.0,
    }
    # whatever a revision titles a column, this is what belongs under it
    by_header = {
        "Peer (auto)": expected["name"], "Price": 142.0,
        "Market Cap": 12_400_000_000, "P/E": 34.0, "EV/EBITDA": 22.0,
        "P/S": 9.1, "Gross Mgn %": 71.0, "Gross margin %": 71.0,
        "Net Mgn %": 6.0, "Net margin %": 6.0, "FCF Yield %": 2.1,
        "NetDebt/EBITDA": 0.4, "Net debt / EBITDA": 0.4,
        "Rev Growth %": 18.0, "Rev Growth % (3Y CAGR)": 18.0,
        # the feeds behind the CAGR — read by nothing
        "Rev LTM (now)": 2037.36, "Rev LTM (-3Y)": 1240.0,
    }
    layouts = [
        # the pre-CAGR revision — raw growth sixth
        ["Peer (auto)", "Price", "Market Cap", "P/E", "EV/EBITDA", "P/S",
         "Rev Growth %", "Gross Mgn %", "Net Mgn %", "FCF Yield %",
         "NetDebt/EBITDA"],
        # what ships today — feeds, then the computed CAGR, at the far right
        ["Peer (auto)", "Price", "Market Cap", "P/E", "EV/EBITDA", "P/S",
         "Gross Mgn %", "Net Mgn %", "FCF Yield %", "NetDebt/EBITDA",
         "Rev LTM (now)", "Rev LTM (-3Y)", "Rev Growth % (3Y CAGR)"],
        # reordered, margins spelled out, the name column no longer first
        ["Rev Growth % (3Y CAGR)", "Net debt / EBITDA", "Peer (auto)",
         "Net margin %", "Gross margin %", "FCF Yield %", "P/S", "EV/EBITDA",
         "P/E", "Market Cap", "Price"],
    ]
    for layout in layouts:
        wb = Workbook()
        ws = wb.active
        ws.title = "Peers"
        for j, title in enumerate(layout, 1):
            ws.cell(row=4, column=j, value=title)
            ws.cell(row=5, column=j, value=by_header[title])
        assert _read_peers(ws) == [expected], layout


def test_valuation_wacc_and_reverse_dcf_keys(workspace):
    data = load_company_data(workspace)
    val = data.valuation
    # the bare 'WACC' row — NOT the 'WACC (auto — CAPM)' band title nor the
    # 'WACC −1%' / 'WACC +1%' sensitivity rows (Unicode-minus labels)
    assert val["wacc"] == 0.092
    # the reverse-DCF row, not the later 'Implied growth sensitivity …' title
    assert val["implied_growth"] == 0.061
    assert val["hist_fcf_cagr"] == 0.045
    assert val["rev_cagr"] == 0.12
    assert val["priced_vs_delivered"] == 0.016
    # a free-text verdict, kept verbatim as a string
    assert val["reverse_dcf_read"] == "market prices in MORE growth than history delivered"
    # the existing bear/base/bull scenario reader is untouched
    assert len(val["scenarios"]) == 3


def test_missing_news_sheet_degrades(workspace):
    # dropping the OPTIONAL News sheet must degrade (news == []), never raise
    from openpyxl import load_workbook

    xlsx = workspace / "dennis_data.xlsx"
    wb = load_workbook(xlsx)
    wb.remove(wb["News"])
    wb.save(xlsx)
    data = load_company_data(workspace)
    assert data.news == []
    # the other optional blocks are unaffected
    assert data.has_peers and len(data.peer_percentiles) == 8


def test_prompt_block_carries_history_and_dashboard(workspace):
    data = load_company_data(workspace)
    block = data.as_prompt_block()
    assert "[history" in block
    assert "revenue:" in block and "FY-0" in block
    assert "[dashboard" in block and "Net margin (LTM):" in block
    # the vendor is never named (only the RIC ticker itself may carry a dot)
    assert "refinitiv" not in block.lower() and "lseg" not in block.lower()


def test_csv_is_snapshot_only(workspace, fixtures_dir):
    (workspace / "dennis_data.xlsx").unlink()
    csv_src = fixtures_dir / "company_data" / "dennis_data.csv"
    (workspace / "dennis_data.csv").write_bytes(csv_src.read_bytes())
    data = load_company_data(workspace)
    assert data.get("ticker") == "EXMPL"
    assert not data.has_history, "CSV carries no history sheet"
    assert data.news == [] and data.peer_percentiles == []


def test_missing_export_raises(settings):
    empty = settings.workspace_dir / "NONE" / "2026-07-01"
    empty.mkdir(parents=True)
    assert find_export(empty) is None
    with pytest.raises(CompanyDataError, match="dennis_data"):
        load_company_data(empty)


def test_na_values_and_unknown_fields(workspace):
    data = load_company_data(workspace)
    assert data.get("pe_ttm") is None       # '#N/A' in the fixture
    assert "pe_ttm" in data.warning_missing


def test_the_value_column_is_found_however_the_template_titles_it(tmp_path):
    """The v3.1 template renamed `Value (auto)` to `Value (Capital IQ) MM`.

    An exact-match column lookup read ZERO snapshot fields off it — a workbook
    that was present, opened cleanly, and parsed to nothing, which then looks
    identical to "the operator never refreshed it". Match on the prefix.
    """
    from openpyxl import Workbook

    for title in ("Value (auto)", "Value", "Value (Capital IQ) MM",
                  "Value (Refinitiv)"):
        wb = Workbook()
        ws = wb.active
        ws.title = "Snapshot"
        ws.cell(row=6, column=2, value="field_key")
        ws.cell(row=6, column=3, value="Label")
        ws.cell(row=6, column=4, value=title)
        for i, (k, v) in enumerate([("company_name", "Example Inc"),
                                    ("ticker", "EXMPL"),
                                    ("as_of_date", "2026-07-01"),
                                    ("price", 12.5),
                                    ("market_cap", 1.0e9),
                                    ("shares_out", 8.0e7)]):
            ws.cell(row=7 + i, column=2, value=k)
            ws.cell(row=7 + i, column=4, value=v)
        d = tmp_path / title.replace(" ", "_").replace("(", "").replace(")", "")
        d.mkdir(parents=True)
        wb.save(d / "dennis_data.xlsx")

        data = load_company_data(d)
        assert data.get("company_name") == "Example Inc", f"{title!r} read nothing"
        assert data.blocking_missing == [], title


def test_the_shipped_template_parses_structurally(tmp_path):
    """Guard the whole contract: whatever revision of the template ships, its
    Snapshot sheet must still be readable by field_key."""
    import shutil
    from pathlib import Path

    template = Path(__file__).resolve().parents[1] / "templates" / "dennis_data_template.xlsx"
    d = tmp_path / "ws"
    d.mkdir(parents=True)
    shutil.copy(template, d / "dennis_data.xlsx")

    # The shipped template holds formulas with no cached values, so every value
    # reads None — but the KEYS must all be found, which is what broke.
    from pipeline.company_data import _read_snapshot
    from openpyxl import load_workbook
    from pipeline.models import ALL_DATA_FIELDS

    wb = load_workbook(d / "dennis_data.xlsx", data_only=True)
    pairs = _read_snapshot(wb["Snapshot"])
    wb.close()
    assert len(pairs) >= 40, f"only {len(pairs)} field_keys found in the template"
    unknown = [k for k in pairs if k not in ALL_DATA_FIELDS]
    assert not unknown, f"template has field_keys the model doesn't know: {unknown}"

    data = load_company_data(d)
    assert data.has_history
    assert data.history_years == ["FY-4", "FY-3", "FY-2", "FY-1", "FY-0", "LTM"]


def test_the_shipped_template_peer_table_maps_every_field():
    """The peer half of the same contract, against the REAL workbook: whatever
    revision ships, every field must find its OWN column, `rev_growth` must be
    the computed 3Y CAGR, and no raw `Rev LTM` figure may reach a metric."""
    from pathlib import Path

    from openpyxl import load_workbook

    from pipeline.company_data import (
        _PEER_COLS,
        _cell_text,
        _col_of,
        _header_row,
        _num,
        _read_peers,
        _rows,
    )

    template = Path(__file__).resolve().parents[1] / "templates" / "dennis_data_template.xlsx"
    wb = load_workbook(template, data_only=True)
    ws = wb["Peers"]
    rows = _rows(ws)
    hr = _header_row(rows, "peer (auto)")
    assert hr is not None, "the auto peer table lost its `Peer (auto)` header"
    header = rows[hr]

    cols = {f: _col_of(header, *names) for f, names in _PEER_COLS.items()}
    assert all(c is not None for c in cols.values()), \
        f"no column found for {[f for f, c in cols.items() if c is None]}"
    # one column each: never two fields on one, never a `cln …` scoring twin
    assert len(set(cols.values())) == len(cols)
    titles = {f: _cell_text(header[c]).lower() for f, c in cols.items()}
    assert not any(t.startswith("cln") for t in titles.values()), titles
    assert "cagr" in titles["rev_growth"] and "ltm" not in titles["rev_growth"]

    peers = _read_peers(ws)
    assert len(peers) > 1 and all(p["name"] for p in peers)
    # per row, the two revenue feeds must not appear in any metric — that is
    # exactly what the fixed-offset read did (`Rev LTM (now)` ≈ 1e5 arriving
    # as a net-debt/EBITDA ratio).
    feed_cols = [c for c in (_col_of(header, "rev ltm (now)"),
                             _col_of(header, "rev ltm (-3y)")) if c is not None]
    assert feed_cols, "the template dropped the revenue feeds"
    for row, peer in zip(rows[hr + 1:], peers):
        fed = {_num(row[c]) for c in feed_cols if c < len(row)} - {None}
        metrics = {v for k, v in peer.items() if k != "name" and v is not None}
        assert not (fed & metrics), peer["name"]
    wb.close()


def test_screenshot_gets_generic_filing_label(workspace, settings, tmp_path):
    src = tmp_path / "raw.png"
    Image.new("RGB", (1400, 800), (30, 34, 40)).save(src)
    out = prepare_screenshot(src, tmp_path / "norm.png", settings)
    img = Image.open(out)
    assert img.size == settings.long_resolution
    assert FILING_LABEL == "FROM THE 10-K"
    # the label chip is drawn in the top-left margin — the corner must no
    # longer be the plain canvas color
    corner = img.crop((0, 0, 200, 60))
    colors = corner.getcolors(maxcolors=4096)
    assert colors and len(colors) > 2, "label chip must be present"


def test_list_screenshots_excludes_exports(workspace):
    (workspace / "income_statement.png").write_bytes(b"png")
    (workspace / "_hidden.png").write_bytes(b"png")
    shots = list_screenshots(workspace)
    assert shots == ["income_statement.png"]
