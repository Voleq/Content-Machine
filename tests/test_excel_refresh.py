"""Driving Excel for the numbers (P3.1b).

There is no Excel on this box, so the COM calls themselves cannot be
exercised here — `Win32ExcelSession` is verified on the Windows machine. What
*is* tested is everything that decides whether the data is trustworthy: the
poll loop that waits for an asynchronous add-in, the classification of
pending vs errored cells, the refusal to publish a half-refreshed workbook,
and the guarantee that the manual upload path is untouched.

The fake add-in below resolves its fields gradually, which is the behaviour
that makes this part dangerous: a naive implementation reads the sheet
immediately, sees blanks, and ships a video full of nothing.
"""

from __future__ import annotations

import json
from datetime import datetime, timedelta, timezone
from pathlib import Path

import pytest
from openpyxl import load_workbook

from config import Settings
from pipeline.company_data import find_export, load_company_data
from pipeline.excel_refresh import (
    FIRST_DATA_ROW,
    KEY_COL,
    PRIORITY_COL,
    RIC_OVERRIDE_CELL,
    TICKER_CELL,
    VALUE_COL,
    ExcelUnavailable,
    RefreshError,
    RefreshTimeout,
    classify_cell,
    excel_available,
    read_snapshot_state,
    refresh_age_days,
    refresh_for_ticker,
    refresh_stamp,
    resolve_ric_override,
    resolve_symbol,
    set_symbol_override,
    template_path,
)
from pipeline.gates import check_freshness
from pipeline.models import DATA_REQUIRED

ROOT = Path(__file__).resolve().parents[1]
TEMPLATE = ROOT / "templates" / "dennis_data_template.xlsx"


# --------------------------------------------------------------------------
# A fake add-in: fields land a few at a time, exactly as CIQ/LSEG behave.
# --------------------------------------------------------------------------


class FakeExcel:
    """Stands in for Excel + the add-in.

    `resolve_after` is how many polls each batch of fields takes to arrive, so
    a test can make the refresh finish promptly, dawdle, or never finish.
    """

    def __init__(self, *, fields: dict[str, object] | None = None,
                 resolve_after: int = 2, macros: tuple[str, ...] = ("EikonRefreshWorksheet",),
                 stuck: tuple[str, ...] = (), errors: tuple[str, ...] = (),
                 priority: dict[str, str] | None = None):
        self.fields = dict(fields or _default_fields())
        # mirrors the v3.1 template: DATA_REQUIRED plus a few the sheet grades
        # Required without them blocking a render
        self.priority = dict(priority if priority is not None else {
            **{k: "Required" for k in DATA_REQUIRED},
            "pe_ttm": "Required",
        })
        self.resolve_after = resolve_after
        self.macros = macros
        self.stuck = set(stuck)
        self.errors = set(errors)
        self.polls = 0
        self.cells: dict[tuple[str, str], object] = {}
        self.opened: Path | None = None
        self.saved: Path | None = None
        self.closed = False
        self.macro_calls: list[str] = []
        self.recalcs = 0
        self.triggered = False

    # ---------------------------------------------------------- the protocol
    def open_workbook(self, path: Path) -> None:
        self.opened = Path(path)

    def set_cell(self, sheet: str, cell: str, value: object) -> None:
        self.cells[(sheet, cell)] = value

    def run_macro(self, name: str) -> None:
        self.macro_calls.append(name)
        if name not in self.macros:
            raise RuntimeError(f"Cannot run the macro '{name}'")
        self.triggered = True

    def calculate(self) -> None:
        self.recalcs += 1
        self.triggered = True

    def calculation_done(self) -> bool:
        return True

    def read_block(self, sheet, first_row, last_row, first_col, last_col):
        self.polls += 1
        rows = []
        for key, value in self.fields.items():
            if key in self.errors:
                shown: object = "#NAME?"
            elif not self.triggered or key in self.stuck:
                shown = "#N/A"
            elif self.polls < self.resolve_after:
                shown = "Requesting Data..."
            else:
                shown = value
            row = [None] * (last_col - first_col + 1)
            row[0] = key
            row[VALUE_COL - KEY_COL] = shown
            # the template's own Required/Recommended/Optional grading (col F)
            pri_i = PRIORITY_COL - KEY_COL
            if pri_i < len(row):
                row[pri_i] = self.priority.get(key, "Optional")
            rows.append(row)
        return rows

    def save_as(self, path: Path) -> None:
        self.saved = Path(path)
        # A real SaveAs writes cached values; the shipped template's formulas
        # would read as None, so bake the resolved values in.
        _write_workbook(self.saved, self.fields)

    def close(self) -> None:
        self.closed = True


def _default_fields() -> dict[str, object]:
    fields: dict[str, object] = {
        "company_name": "Palantir Technologies Inc",
        "ticker": "PLTR.O",
        "as_of_date": datetime(2026, 7, 24),
        "price": 158.4,
        "market_cap": 372_000.0,
        "shares_out": 2_350.0,
        "revenue_ltm": 3_120.0,
        "pe_ttm": 214.0,
    }
    return fields


def _write_workbook(path: Path, fields: dict[str, object]) -> None:
    """A minimal Snapshot sheet in the v3.1 layout the reader expects.

    Note the value column header: the real template calls it
    `Value (Capital IQ) MM`, and an exact-match column lookup read ZERO fields
    off it. Mirroring the real title keeps that regression caught here.
    """
    from openpyxl import Workbook

    header = FIRST_DATA_ROW - 1
    wb = Workbook()
    ws = wb.active
    ws.title = "Snapshot"
    ws.cell(row=header, column=1, value="Section")
    ws.cell(row=header, column=KEY_COL, value="field_key")
    ws.cell(row=header, column=3, value="Label")
    ws.cell(row=header, column=VALUE_COL, value="Value (Capital IQ) MM")
    ws.cell(row=header, column=PRIORITY_COL, value="Priority")
    for i, (key, value) in enumerate(fields.items()):
        r = FIRST_DATA_ROW + i
        ws.cell(row=r, column=KEY_COL, value=key)
        ws.cell(row=r, column=VALUE_COL, value=value)
        ws.cell(row=r, column=PRIORITY_COL,
                value="Required" if key in DATA_REQUIRED else "Optional")
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


@pytest.fixture()
def no_sleep(monkeypatch):
    """A clock that advances only when the code sleeps — no wall time."""
    state = {"t": 0.0}

    def sleep(sec: float) -> None:
        state["t"] += max(sec, 0.01)

    return sleep, (lambda: state["t"])


# --------------------------------------------------------------------------
# Cell classification — the difference between "wait" and "wrong".
# --------------------------------------------------------------------------


def test_pending_markers_are_not_mistaken_for_data():
    for pending in ("Requesting Data...", "#N/A", "  loading ", "#N/A N/A",
                    "Retrieving...", ""):
        assert classify_cell(pending) == "pending", pending


def test_a_missing_addin_reads_as_an_error_not_a_wait():
    """`#NAME?` means the TR function itself is unknown — waiting won't fix it."""
    assert classify_cell("#NAME?") == "error"
    assert classify_cell("#VALUE!") == "error"


def test_real_values_pass_including_zero_and_dates():
    assert classify_cell(0) == "ok"          # a genuine zero is data
    assert classify_cell(158.4) == "ok"
    assert classify_cell("Palantir Technologies Inc") == "ok"
    assert classify_cell(datetime(2026, 7, 24)) == "ok"


def test_snapshot_state_reads_keys_and_values_off_the_block():
    rows = [["price", None, 158.4], ["market_cap", None, "#N/A"], [None, None, None]]
    st = read_snapshot_state(rows)
    assert st.values["price"] == 158.4
    assert st.pending == ["market_cap"]
    assert st.resolved == 1
    assert st.unresolved_required(["price", "market_cap"]) == ["market_cap"]


# --------------------------------------------------------------------------
# The poll loop — the part that keeps empty data out.
# --------------------------------------------------------------------------


def test_a_refresh_waits_for_the_addin_before_saving(settings, no_sleep, tmp_path):
    """The whole point: the sheet is read repeatedly, not once."""
    sleep, clock = no_sleep
    fake = FakeExcel(resolve_after=4)
    ws_dir = tmp_path / "ws"
    result = refresh_for_ticker(settings, "PLTR", ws_dir, session_factory=lambda: fake,
                               sleep=sleep, clock=clock)
    assert fake.polls >= 4, "gave up before the add-in had answered"
    assert result.resolved == len(_default_fields())
    assert not result.pending


def test_it_keeps_polling_until_the_picture_stops_changing(settings, no_sleep, tmp_path):
    """Required fields landing is not the same as the refresh being done.

    Stopping on the first complete read would save while the rest of the
    workbook is still filling in, so the loop also waits for N consecutive
    unchanged reads. Asserted differentially: raising the settle count has to
    cost more polls, which only holds if the condition is really applied.
    """
    sleep, clock = no_sleep
    counts = {}
    for settle in (1, 5):
        fake = FakeExcel(resolve_after=3)
        s = settings.model_copy(update={"excel_settle_polls": settle})
        refresh_for_ticker(s, "PLTR", tmp_path / f"ws{settle}", session_factory=lambda: fake,
                           sleep=sleep, clock=clock)
        counts[settle] = fake.polls
    # 2 pending polls, the poll that resolves, then `settle` unchanged ones
    assert counts[1] == 4
    assert counts[5] == 8


def test_a_timeout_is_a_hard_failure_and_writes_nothing(settings, no_sleep, tmp_path):
    """An add-in that never answers must not yield a usable-looking workbook."""
    sleep, clock = no_sleep
    fake = FakeExcel(stuck=("price", "market_cap"))
    ws_dir = tmp_path / "ws"
    with pytest.raises(RefreshTimeout) as e:
        refresh_for_ticker(settings, "PLTR", ws_dir, session_factory=lambda: fake,
                           sleep=sleep, clock=clock, timeout_s=20)
    assert "price" in str(e.value) and "market_cap" in str(e.value)
    # the symbol we typed is named too — a wrong vendor code is the commonest
    # cause and is invisible from the field list alone
    assert "PLTR" in str(e.value)
    assert find_export(ws_dir) is None, "a failed refresh left data behind"
    assert not list(ws_dir.glob("*.xlsx"))


def test_a_stale_workbook_survives_a_failed_refresh(settings, no_sleep, tmp_path):
    """Failure changes nothing — the previous numbers stay exactly as they were."""
    sleep, clock = no_sleep
    ws_dir = tmp_path / "ws"
    ws_dir.mkdir(parents=True)
    existing = ws_dir / "dennis_data.xlsx"
    _write_workbook(existing, {"company_name": "Old Co", "price": 1.0})
    before = existing.read_bytes()

    with pytest.raises(RefreshTimeout):
        refresh_for_ticker(settings, "PLTR", ws_dir,
                           session_factory=lambda: FakeExcel(stuck=("price",)),
                           sleep=sleep, clock=clock, timeout_s=20)
    assert existing.read_bytes() == before


def test_an_unloaded_addin_fails_with_the_field_named(settings, no_sleep, tmp_path):
    """`#NAME?` everywhere means the add-in isn't loaded. Say so, don't hang."""
    sleep, clock = no_sleep
    fake = FakeExcel(errors=("price", "market_cap"))
    with pytest.raises(RefreshTimeout) as e:
        refresh_for_ticker(settings, "PLTR", tmp_path / "ws", session_factory=lambda: fake,
                           sleep=sleep, clock=clock, timeout_s=15)
    msg = str(e.value)
    assert "price" in msg
    assert "errored" in msg


def test_optional_fields_that_never_resolve_do_not_fail_the_refresh(
        settings, no_sleep, tmp_path):
    """A missing P/E is a warning; a missing price is a failure."""
    sleep, clock = no_sleep
    fake = FakeExcel(stuck=("pe_ttm",))
    result = refresh_for_ticker(settings, "PLTR", tmp_path / "ws", session_factory=lambda: fake,
                               sleep=sleep, clock=clock, timeout_s=30)
    assert "pe_ttm" in result.pending
    assert result.path.exists()
    assert all(f not in result.pending for f in DATA_REQUIRED)


def test_the_poll_waits_for_the_templates_required_set_not_just_the_hard_one(
        settings, no_sleep, tmp_path):
    """`pe_ttm` does not block a render, but the poll should still wait for it:
    stopping earlier risks saving while the valuation block fills in.

    Isolated with a field that NEVER lands — one that eventually lands is
    already covered by the settle condition, so it cannot tell the two grades
    apart. Graded Required, the poll holds out to the deadline; graded
    Optional, it settles and leaves.
    """
    sleep, clock = no_sleep
    counts = {}
    for grade in ("Required", "Optional"):
        fake = FakeExcel(priority={**{k: "Required" for k in DATA_REQUIRED},
                                   "pe_ttm": grade},
                         stuck=("pe_ttm",))
        refresh_for_ticker(settings, "PLTR", tmp_path / f"ws{grade}",
                           session_factory=lambda f=fake: f,
                           sleep=sleep, clock=clock, timeout_s=30)
        counts[grade] = fake.polls
    assert counts["Optional"] == 5, counts     # resolves at 2, settles by 5
    assert counts["Required"] > counts["Optional"], counts


def test_a_template_required_field_that_never_lands_warns_but_still_saves(
        settings, no_sleep, tmp_path):
    """The soft tier is a warning, not a failure: a thin small-cap missing
    ev_ebitda should still be able to get a video."""
    sleep, clock = no_sleep
    fake = FakeExcel(priority={**{k: "Required" for k in DATA_REQUIRED},
                               "pe_ttm": "Required"},
                     stuck=("pe_ttm",))
    result = refresh_for_ticker(settings, "PLTR", tmp_path / "ws",
                                session_factory=lambda: fake,
                                sleep=sleep, clock=clock, timeout_s=30)
    assert result.path.exists(), "a soft-tier gap must not fail the refresh"
    assert "pe_ttm" in result.pending
    assert "pe_ttm" in result.summary()


# --------------------------------------------------------------------------
# Triggering the refresh.
# --------------------------------------------------------------------------


def test_the_first_macro_that_exists_wins(settings, no_sleep, tmp_path):
    sleep, clock = no_sleep
    fake = FakeExcel(macros=("SPRefreshAll",))
    result = refresh_for_ticker(settings, "PLTR", tmp_path / "ws", session_factory=lambda: fake,
                               sleep=sleep, clock=clock)
    assert result.macro == "SPRefreshAll"
    # it tried the earlier candidates and moved on rather than giving up
    assert fake.macro_calls[0] != "SPRefreshAll"
    assert fake.recalcs == 0


def test_no_macro_falls_back_to_a_full_recalculation(settings, no_sleep, tmp_path):
    """An add-in whose macro name we don't know still refreshes — its formulas
    are volatile, so a rebuild re-issues them."""
    sleep, clock = no_sleep
    fake = FakeExcel(macros=())
    result = refresh_for_ticker(settings, "PLTR", tmp_path / "ws", session_factory=lambda: fake,
                               sleep=sleep, clock=clock)
    assert result.macro == ""
    assert fake.recalcs == 1
    assert result.resolved == len(_default_fields())


def test_the_macro_list_is_configurable(settings, no_sleep, tmp_path):
    sleep, clock = no_sleep
    fake = FakeExcel(macros=("MyOwnRefresh",))
    s = settings.model_copy(update={"excel_refresh_macros": " MyOwnRefresh , Other "})
    result = refresh_for_ticker(s, "PLTR", tmp_path / "ws", session_factory=lambda: fake,
                               sleep=sleep, clock=clock)
    assert result.macro == "MyOwnRefresh"
    assert fake.macro_calls == ["MyOwnRefresh"]


# --------------------------------------------------------------------------
# The two input cells. The v3.1 template takes the plain ticker in C3 and
# derives the Refinitiv RIC itself in B3; E2 forces a RIC when the derivation
# cannot know better. Writing the RIC into C3 would break every CIQ formula.
# --------------------------------------------------------------------------


def test_the_plain_ticker_lands_in_the_ticker_cell(settings, no_sleep, tmp_path):
    sleep, clock = no_sleep
    fake = FakeExcel()
    refresh_for_ticker(settings, "pltr", tmp_path / "ws",
                       session_factory=lambda: fake, sleep=sleep, clock=clock)
    assert TICKER_CELL == "C3"
    assert fake.cells[("Snapshot", TICKER_CELL)] == "PLTR"
    # nothing forced, so the template derives the RIC from the exchange
    assert ("Snapshot", RIC_OVERRIDE_CELL) not in fake.cells


def test_the_template_really_does_read_those_cells():
    """Guard against the template moving under us.

    Both halves matter: C3 is what the CIQ formulas consume, and B3 must stay
    a formula — writing a RIC into it would silently detach every green cell
    from the ticker.
    """
    wb = load_workbook(TEMPLATE)
    ws = wb["Snapshot"]
    formulas = [str(c.value) for row in ws.iter_rows() for c in row
                if isinstance(c.value, str) and c.value.startswith("=")]
    assert any(f"CIQ({TICKER_CELL}" in f.replace(" ", "") for f in formulas), \
        "no CIQ formula reads the ticker cell"
    derived = str(ws["B3"].value)
    assert derived.startswith("="), "B3 should be the derived RIC, not an input"
    assert RIC_OVERRIDE_CELL in derived.replace("$", ""), \
        "B3 no longer honours the RIC override cell"
    wb.close()


def test_the_template_grades_its_own_required_fields():
    """The Priority column is what the poll waits for."""
    wb = load_workbook(TEMPLATE)
    ws = wb["Snapshot"]
    graded = {str(r[KEY_COL - 1].value).strip(): str(r[PRIORITY_COL - 1].value).strip()
              for r in ws.iter_rows(min_row=FIRST_DATA_ROW, max_col=PRIORITY_COL)
              if r[KEY_COL - 1].value}
    required = {k for k, p in graded.items() if p == "Required"}
    assert set(DATA_REQUIRED) <= required, \
        "the template no longer grades every blocking field as Required"
    assert len(required) > len(DATA_REQUIRED), \
        "the two tiers have collapsed — the soft tier buys nothing"
    wb.close()


def test_a_pinned_ric_goes_to_the_override_cell_not_the_ticker_cell(
        settings, no_sleep, tmp_path):
    """The pin is a RIC. Typing it into C3 would break every CIQ formula."""
    set_symbol_override(settings, "PLTR", "PLTR.OQ")
    sleep, clock = no_sleep
    fake = FakeExcel()
    result = refresh_for_ticker(settings, "PLTR", tmp_path / "ws",
                                session_factory=lambda: fake,
                                sleep=sleep, clock=clock)
    assert fake.cells[("Snapshot", TICKER_CELL)] == "PLTR"
    assert fake.cells[("Snapshot", RIC_OVERRIDE_CELL)] == "PLTR.OQ"
    assert result.symbol == "PLTR" and result.ric == "PLTR.OQ"


def test_the_ticker_cell_never_receives_a_ric(settings):
    assert resolve_symbol(settings, "pltr") == "PLTR"
    set_symbol_override(settings, "PLTR", "PLTR.O")
    assert resolve_symbol(settings, "PLTR") == "PLTR", "a RIC leaked into C3"
    assert resolve_ric_override(settings, "PLTR") == "PLTR.O"


def test_no_override_by_default_so_the_template_derives_the_ric(settings):
    """_RICMap gets the common exchanges right; don't second-guess it."""
    assert resolve_ric_override(settings, "PLTR") == ""


def test_a_configured_suffix_forces_a_ric(settings):
    s = settings.model_copy(update={"excel_symbol_suffix": "O"})
    assert resolve_ric_override(s, "PLTR") == "PLTR.O"
    # a ticker that already carries a suffix is left to the template
    assert resolve_ric_override(s, "BRK.B") == ""


def test_an_explicit_symbol_forces_the_ric(settings, no_sleep, tmp_path):
    sleep, clock = no_sleep
    fake = FakeExcel()
    result = refresh_for_ticker(settings, "PLTR", tmp_path / "ws",
                                symbol="PLTR.OQ", session_factory=lambda: fake,
                                sleep=sleep, clock=clock)
    assert result.symbol == "PLTR"
    assert result.ric == "PLTR.OQ"
    assert fake.cells[("Snapshot", TICKER_CELL)] == "PLTR"
    assert fake.cells[("Snapshot", RIC_OVERRIDE_CELL)] == "PLTR.OQ"
    assert "PLTR.OQ" in result.summary()


# --------------------------------------------------------------------------
# What lands on disk.
# --------------------------------------------------------------------------


def test_the_refreshed_workbook_loads_through_the_existing_reader(
        settings, no_sleep, tmp_path):
    """The point of the dated copy: the reader is untouched by any of this."""
    sleep, clock = no_sleep
    ws_dir = tmp_path / "ws"
    result = refresh_for_ticker(settings, "PLTR", ws_dir, session_factory=FakeExcel,
                               sleep=sleep, clock=clock)
    assert find_export(ws_dir) == ws_dir / "dennis_data.xlsx"
    data = load_company_data(ws_dir)
    assert data.get("company_name") == "Palantir Technologies Inc"
    assert data.get("price") == 158.4
    assert not data.blocking_missing
    # the dated archive stays alongside it
    assert result.archive.exists()
    assert "PLTR" in result.archive.name and result.archive != result.path


def test_the_template_is_never_written_to(settings, no_sleep, tmp_path):
    sleep, clock = no_sleep
    before = TEMPLATE.read_bytes()
    fake = FakeExcel()
    refresh_for_ticker(settings, "PLTR", tmp_path / "ws", session_factory=lambda: fake,
                       sleep=sleep, clock=clock)
    assert TEMPLATE.read_bytes() == before
    assert fake.opened != template_path(settings)
    assert fake.opened is not None and fake.opened.parent == tmp_path / "ws"


def test_no_scratch_files_are_left_in_the_workspace(settings, no_sleep, tmp_path):
    sleep, clock = no_sleep
    ws_dir = tmp_path / "ws"
    refresh_for_ticker(settings, "PLTR", ws_dir, session_factory=FakeExcel,
                       sleep=sleep, clock=clock)
    assert not list(ws_dir.glob(".refresh_*"))


def test_an_older_csv_export_cannot_outlive_the_refresh(settings, no_sleep, tmp_path):
    """`find_export` prefers .xlsx, but a leftover CSV is a trap for a reader
    that ever changes its mind about precedence."""
    sleep, clock = no_sleep
    ws_dir = tmp_path / "ws"
    ws_dir.mkdir(parents=True)
    (ws_dir / "dennis_data.csv").write_text("field_key,value\nprice,1.0\n")
    refresh_for_ticker(settings, "PLTR", ws_dir, session_factory=FakeExcel,
                       sleep=sleep, clock=clock)
    assert not (ws_dir / "dennis_data.csv").exists()
    assert (ws_dir / "dennis_data.superseded.csv").exists()


def test_excel_is_closed_on_the_happy_path(settings, no_sleep, tmp_path):
    """A stray Excel process is the failure mode that ruins the next render."""
    sleep, clock = no_sleep
    fake = FakeExcel()
    refresh_for_ticker(settings, "PLTR", tmp_path / "ws",
                       session_factory=lambda: fake, sleep=sleep, clock=clock)
    assert fake.closed


@pytest.mark.parametrize("break_at", ["open_workbook", "set_cell", "save_as",
                                      "read_block"])
def test_excel_is_closed_however_the_refresh_fails(settings, no_sleep, tmp_path,
                                                   break_at):
    """Every failure path, not just the tidy one."""
    sleep, clock = no_sleep
    fake = FakeExcel()

    def explode(*a, **k):
        raise RuntimeError(f"COM blew up in {break_at}")

    setattr(fake, break_at, explode)
    with pytest.raises(Exception):
        refresh_for_ticker(settings, "PLTR", tmp_path / "ws",
                           session_factory=lambda: fake, sleep=sleep, clock=clock,
                           timeout_s=10)
    assert fake.closed, f"{break_at} failure left Excel running"
    assert not list((tmp_path / "ws").glob(".refresh_*"))


def test_a_teardown_failure_does_not_swallow_the_real_error(settings, no_sleep,
                                                            tmp_path):
    """If Quit() also fails, the operator still needs to hear about the timeout."""
    sleep, clock = no_sleep

    class WontClose(FakeExcel):
        def close(self):
            raise RuntimeError("Excel ignored Quit()")

    with pytest.raises(RefreshTimeout):
        refresh_for_ticker(settings, "PLTR", tmp_path / "ws",
                           session_factory=lambda: WontClose(stuck=("price",)),
                           sleep=sleep, clock=clock, timeout_s=10)


def test_a_missing_template_is_reported_not_crashed_on(settings, tmp_path):
    s = settings.model_copy(update={"excel_template_path": str(tmp_path / "nope.xlsx")})
    with pytest.raises(ExcelUnavailable) as e:
        refresh_for_ticker(s, "PLTR", tmp_path / "ws", session_factory=FakeExcel)
    assert "template" in str(e.value)


def test_a_session_that_writes_no_file_is_treated_as_failure(settings, no_sleep, tmp_path):
    """Trust the disk, not Excel's word for it."""
    sleep, clock = no_sleep

    class SilentlyFails(FakeExcel):
        def save_as(self, path):
            self.saved = Path(path)      # says yes, writes nothing

    with pytest.raises(RefreshError) as e:
        refresh_for_ticker(settings, "PLTR", tmp_path / "ws",
                           session_factory=SilentlyFails, sleep=sleep, clock=clock)
    assert "wrote no file" in str(e.value)


# --------------------------------------------------------------------------
# Availability, and the manual path that must keep working.
# --------------------------------------------------------------------------


def test_excel_is_unavailable_on_this_host_and_says_why(settings):
    ok, why = excel_available(settings)
    import sys
    if sys.platform != "win32":
        assert not ok
        assert "Windows" in why


def test_the_switch_turns_it_off_without_touching_anything_else(settings):
    s = settings.model_copy(update={"excel_refresh_enabled": False})
    ok, why = excel_available(s)
    assert not ok
    assert "switched off" in why


def test_the_manual_upload_path_is_unchanged(settings, tmp_path):
    """The fallback is the whole safety net: no COM, no refresh stamp, still
    loads."""
    ws_dir = tmp_path / "ws"
    _write_workbook(ws_dir / "dennis_data.xlsx", _default_fields())
    data = load_company_data(ws_dir)
    assert data.get("price") == 158.4
    assert refresh_stamp(ws_dir) == {}


# --------------------------------------------------------------------------
# Freshness from the refresh timestamp, not the file.
# --------------------------------------------------------------------------


def test_the_refresh_is_stamped_with_when_it_finished(settings, no_sleep, tmp_path):
    sleep, clock = no_sleep
    ws_dir = tmp_path / "ws"
    result = refresh_for_ticker(settings, "PLTR", ws_dir, session_factory=FakeExcel,
                               sleep=sleep, clock=clock)
    stamp = refresh_stamp(ws_dir)
    assert stamp["symbol"] == "PLTR"
    assert stamp["source"] == "excel_com"
    assert stamp["resolved_fields"] == result.resolved
    assert datetime.fromisoformat(stamp["finished_at"]).tzinfo is not None
    assert refresh_age_days(ws_dir) < 1


def test_freshness_uses_the_refresh_timestamp_over_the_sheets_as_of_date(
        settings, tmp_path):
    """A workbook re-saved today whose numbers are three weeks old is stale;
    `=TODAY()` would call it fresh."""
    ws_dir = tmp_path / "ws"
    ws_dir.mkdir(parents=True)
    old = datetime.now(timezone.utc) - timedelta(days=30)
    (ws_dir / "data_refresh.json").write_text(json.dumps(
        {"finished_at": old.isoformat(), "symbol": "PLTR"}))

    today = datetime.now(timezone.utc).date().isoformat()
    assert check_freshness(today, settings) == []          # sheet says today
    findings = check_freshness(today, settings, workspace=ws_dir)
    assert len(findings) == 1
    assert "refreshed 30" in findings[0].message
    assert "/refresh" in findings[0].message


def test_a_recent_refresh_is_silent(settings, tmp_path):
    ws_dir = tmp_path / "ws"
    ws_dir.mkdir(parents=True)
    (ws_dir / "data_refresh.json").write_text(json.dumps(
        {"finished_at": datetime.now(timezone.utc).isoformat()}))
    assert check_freshness("", settings, workspace=ws_dir) == []


def test_without_a_stamp_freshness_falls_back_to_the_as_of_date(settings, tmp_path):
    """Manually uploaded workbooks have no stamp — the old check still runs."""
    ws_dir = tmp_path / "ws"
    ws_dir.mkdir(parents=True)
    stale = (datetime.now(timezone.utc) - timedelta(days=40)).date().isoformat()
    findings = check_freshness(stale, settings, workspace=ws_dir)
    assert len(findings) == 1
    assert "days old" in findings[0].message


def test_stale_data_can_be_made_blocking(settings, tmp_path):
    ws_dir = tmp_path / "ws"
    ws_dir.mkdir(parents=True)
    old = datetime.now(timezone.utc) - timedelta(days=30)
    (ws_dir / "data_refresh.json").write_text(json.dumps(
        {"finished_at": old.isoformat()}))
    s = settings.model_copy(update={"data_stale_blocks": True})
    assert check_freshness("", s, workspace=ws_dir)[0].severity == "block"


# --------------------------------------------------------------------------
# The bot side.
# --------------------------------------------------------------------------


def test_the_bot_offers_the_template_when_excel_is_absent(settings, tmp_path):
    """On this Linux host /refresh must degrade to the manual instruction."""
    from bot.handlers import BotCore

    core = BotCore(settings)
    core.new_ticker(11, "PLTR")
    reply = core.refresh_data(11, ["PLTR"])
    assert "Can't drive Excel" in reply.text
    assert "unchanged" in reply.text
    assert any(f.name.endswith(".xlsx") for f in reply.files)


def test_new_still_sends_the_template_when_it_cannot_refresh(settings):
    from bot.handlers import BotCore
    import sys

    core = BotCore(settings)
    reply = core.new_ticker(12, "PLTR")
    if sys.platform != "win32":
        assert reply.files, "no template attached and no way to refresh"


def test_refresh_reports_the_failure_and_keeps_the_chat_usable(
        settings, monkeypatch, tmp_path):
    from bot import handlers

    core = handlers.BotCore(settings)
    core.new_ticker(13, "PLTR")
    monkeypatch.setattr(handlers, "excel_available", lambda s: (True, "ok"))

    def boom(*a, **k):
        raise RefreshTimeout("required field(s) still empty: price")

    monkeypatch.setattr(handlers, "refresh_for_ticker", boom)
    reply = core.refresh_data(13, ["PLTR"])
    assert "price" in reply.text
    assert "Nothing was saved" in reply.text


def test_a_successful_refresh_hands_back_the_prompts(settings, monkeypatch):
    """The manual data step disappears: refresh straight into the prompts."""
    from bot import handlers

    core = handlers.BotCore(settings)
    core.new_ticker(14, "PLTR")
    ws = core.context.get(14)

    monkeypatch.setattr(handlers, "excel_available", lambda s: (True, "ok"))
    monkeypatch.setattr(handlers, "refresh_for_ticker", _fake_refresh)
    reply = core.refresh_data(14, ["PLTR"])

    assert "Refreshed PLTR" in reply.text
    names = [f.name for f in reply.files]
    assert any(n.startswith("prompt_") for n in names), names
    assert any(n.startswith("dennis_data_PLTR_") for n in names), names
    assert (ws.path / "dennis_data.xlsx").exists()


def test_the_chat_summary_never_names_the_data_vendor(settings, no_sleep, tmp_path):
    """The macro name is the vendor's brand. It goes in the log and the stamp
    file, never in a message."""
    sleep, clock = no_sleep
    fake = FakeExcel(macros=("EikonRefreshWorksheet",))
    result = refresh_for_ticker(settings, "PLTR", tmp_path / "ws",
                                session_factory=lambda: fake,
                                sleep=sleep, clock=clock)
    text = result.summary()
    for brand in ("Eikon", "Refinitiv", "LSEG", "Capital IQ", "CIQ", "Thomson"):
        assert brand not in text, brand
    # but it is recorded where you would go to configure it
    assert refresh_stamp(tmp_path / "ws")["macro"] == "EikonRefreshWorksheet"


def test_the_prompts_reply_says_how_old_the_numbers_really_are(settings, tmp_path):
    """The sheet's `as_of_date` is =TODAY(); the operator needs the truth."""
    from bot.handlers import BotCore

    core = BotCore(settings)
    core.new_ticker(16, "PLTR")
    ws_dir = settings.workspace_dir / "PLTR"
    day = sorted(p for p in ws_dir.iterdir() if p.is_dir())[-1]
    _write_workbook(day / "dennis_data.xlsx", _default_fields())
    old = datetime.now(timezone.utc) - timedelta(days=4)
    (day / "data_refresh.json").write_text(json.dumps(
        {"finished_at": old.isoformat()}))

    reply = core.prompts_reply(16)
    assert "refreshed 4.0 days ago" in reply.text


def test_a_refresh_withdraws_a_pending_approval(settings, monkeypatch,
                                                short_valid_json):
    """Approve → refresh → render would otherwise ship numbers the operator
    never saw: the approval pins the *script* hash, which a data refresh does
    not change."""
    from bot import handlers

    core = handlers.BotCore(settings)
    core.new_ticker(17, "EXMPL")
    ws = core.context.get(17)
    _write_workbook(ws.path / "dennis_data.xlsx", _default_fields())
    core.intake_script(17, short_valid_json)
    script = ws.load_short()
    ws.approve("short", script.content_sha(), "the report they read")
    assert ws.is_approved("short")

    monkeypatch.setattr(handlers, "excel_available", lambda s: (True, "ok"))
    monkeypatch.setattr(handlers, "refresh_for_ticker", _fake_refresh)
    reply = core.refresh_data(17, ["EXMPL"])

    assert not ws.is_approved("short"), "rendered on an approval of stale numbers"
    assert "approval was withdrawn" in reply.text


def _fake_refresh(s, ticker, workspace, **kw):
    from pipeline.excel_refresh import RefreshResult

    _write_workbook(Path(workspace) / "dennis_data.xlsx", _default_fields())
    archive = Path(workspace) / f"dennis_data_{ticker}_2026-07-26.xlsx"
    _write_workbook(archive, _default_fields())
    now = datetime.now(timezone.utc)
    return RefreshResult(path=Path(workspace) / "dennis_data.xlsx",
                         archive=archive, symbol=ticker,
                         started_at=now, finished_at=now, polls=4,
                         macro="EikonRefreshWorksheet", resolved=8)


def test_an_explicit_symbol_on_the_command_is_remembered(settings, monkeypatch):
    """Typing the RIC once should be enough — and it is remembered as a RIC
    override, not as a replacement ticker."""
    from bot import handlers

    core = handlers.BotCore(settings)
    monkeypatch.setattr(handlers, "excel_available", lambda s: (True, "ok"))
    monkeypatch.setattr(handlers, "refresh_for_ticker",
                        lambda *a, **k: (_ for _ in ()).throw(RefreshError("stop")))
    core.refresh_data(15, ["PLTR", "PLTR.OQ"])
    assert resolve_ric_override(settings, "PLTR") == "PLTR.OQ"
    assert resolve_symbol(settings, "PLTR") == "PLTR"
