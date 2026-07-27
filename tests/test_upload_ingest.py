"""The manual-upload data path — the primary way numbers reach the bot.

Excel is refreshed outside the bot (a values-only paste into a clean
workbook) and the file is dropped into the workspace. That makes this the
front door rather than a fallback, so it has to fail like a front door: the
message names the thing the operator can go and fix.

Five failures each have a different fix, and telling them apart is the whole
point of these tests:

    missing sheet   -> re-export, the template lost a tab
    formulas only   -> paste as values; add-in formulas do not travel
    unresolved      -> sign the terminal in and refresh, then re-export
    wrong ticker    -> that is a different company
    stale           -> those numbers are from last month

The sixth property is that a bad upload never destroys a good workbook.
"""

from __future__ import annotations

import datetime as dt
import shutil
from pathlib import Path

import pytest
from openpyxl import load_workbook

from pipeline.company_data import (
    check_export,
    load_company_data,
    unresolved_marker,
)

FIXTURE = Path(__file__).resolve().parents[1] / "fixtures" / "company_data" / "dennis_data.xlsx"

GOOD_ROWS = [
    ("company_name", "Example Industries"),
    ("ticker", "EXMPL"),
    ("as_of_date", "2026-07-20"),
    ("price", 158.4),
    ("market_cap", 12_400_000_000.0),
    ("shares_out", 78_200_000.0),
]


def _write(path: Path, rows, *, sheet: str = "Snapshot",
           value_header: str = "Value (auto)") -> Path:
    """A minimal values-only workbook in the template's Snapshot shape."""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = sheet
    ws.append(["", "field_key", "", value_header, "", "Priority"])
    for key, value in rows:
        ws.append(["", key, "", value, "", "Required"])
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return path


# --------------------------------------------------------------------------
# A values-only workbook is the expected input
# --------------------------------------------------------------------------


def test_a_values_only_workbook_reads_cleanly(tmp_path):
    """No formulas anywhere — the reader opens data_only=True, so this is the
    shape it is built for."""
    src = _write(tmp_path / "dennis_data.xlsx", GOOD_ROWS)
    check = check_export(src, expect_ticker="EXMPL", max_age_days=10,
                         today=dt.date(2026, 7, 22))
    assert check.ok, check.render()
    assert check.problems == []
    assert check.ticker == "EXMPL"
    assert check.as_of == "2026-07-20"

    data = load_company_data(tmp_path)
    assert data.get("company_name") == "Example Industries"
    assert data.get("price") == 158.4
    assert data.blocking_missing == []


def test_the_shipped_fixture_is_a_clean_values_only_export(tmp_path):
    shutil.copy(FIXTURE, tmp_path / "dennis_data.xlsx")
    check = check_export(tmp_path / "dennis_data.xlsx")
    assert check.ok, check.render()

    # ...and it genuinely carries no formulas: every cell is a literal.
    wb = load_workbook(tmp_path / "dennis_data.xlsx", data_only=False)
    formulas = [c.value for sheet in wb.sheetnames for row in wb[sheet].iter_rows()
                for c in row if isinstance(c.value, str) and c.value.startswith("=")]
    wb.close()
    assert formulas == [], f"fixture is not values-only: {formulas[:5]}"


def test_the_value_column_title_may_vary(tmp_path):
    """Template revisions renamed it; a prefix match is why that stopped
    silently reading zero fields."""
    src = _write(tmp_path / "dennis_data.xlsx", GOOD_ROWS,
                 value_header="Value (Capital IQ) MM")
    assert check_export(src).ok


# --------------------------------------------------------------------------
# Unresolved add-in cells
# --------------------------------------------------------------------------


@pytest.mark.parametrize("marker", [
    "#CIQINACTIVE",
    "Not Signed In",
    "#NAME?",
    "Requesting Data...",
    "#VALUE!",
    "#REF!",
])
def test_unresolved_markers_are_recognised(marker):
    assert unresolved_marker(marker) is not None
    assert unresolved_marker(f"  {marker.lower()}  ") is not None


@pytest.mark.parametrize("fine", [
    None, 0, 0.0, 158.4, "Example Industries", "EXMPL",
    dt.date(2026, 7, 20), "#N/A",
])
def test_real_values_are_not_mistaken_for_failures(fine):
    """`#N/A` is deliberately allowed: on a values-only export it is the
    ordinary way a mnemonic says 'no figure for this company', and treating it
    as an alarm would reject good workbooks for thinly-covered small-caps."""
    assert unresolved_marker(fine) is None


def test_an_unresolved_text_field_never_reaches_the_script(tmp_path):
    """The bug this exists for: `company_name` is a string field, so
    `#CIQINACTIVE` satisfied the required-field check and would have been
    spoken and captioned. A video titled `#CIQINACTIVE`."""
    _write(tmp_path / "dennis_data.xlsx", [
        ("company_name", "#CIQINACTIVE"),
        ("ticker", "Not Signed In"),
        ("as_of_date", "2026-07-20"),
        ("price", 158.4),
        ("market_cap", 12_400_000_000.0),
        ("shares_out", 78_200_000.0),
    ])
    data = load_company_data(tmp_path)
    assert data.get("company_name") is None
    assert data.get("ticker") is None
    assert "company_name" in data.blocking_missing
    assert "ticker" in data.blocking_missing


def test_unresolved_required_fields_block_with_a_specific_message(tmp_path):
    src = _write(tmp_path / "dennis_data.xlsx",
                 [("company_name", "#CIQINACTIVE"), ("ticker", "EXMPL"),
                  ("as_of_date", "2026-07-20"), ("price", 158.4),
                  ("market_cap", 1.0), ("shares_out", 1.0)])
    check = check_export(src)
    assert not check.ok
    problem = check.blocking[0]
    assert problem.kind == "unresolved"
    assert "company_name" in problem.message
    assert "#CIQINACTIVE" in problem.message, "show what is actually in the cell"
    assert "signed in" in problem.message, "name the fix"


def test_unresolved_optional_fields_only_warn(tmp_path):
    """A gap in an optional field is not a reason to refuse the upload."""
    src = _write(tmp_path / "dennis_data.xlsx",
                 GOOD_ROWS + [("beta", "#CIQINACTIVE"), ("peg", "Not Signed In")])
    check = check_export(src)
    assert check.ok, check.render()
    assert len(check.warnings) == 1
    assert "beta" in check.warnings[0].message


# --------------------------------------------------------------------------
# Missing sheet
# --------------------------------------------------------------------------


def test_a_missing_snapshot_sheet_says_so(tmp_path):
    src = _write(tmp_path / "dennis_data.xlsx", GOOD_ROWS, sheet="Instructions")
    check = check_export(src)
    assert not check.ok
    problem = check.blocking[0]
    assert problem.kind == "missing_sheet"
    assert "Snapshot" in problem.message
    assert "Instructions" in problem.message, "say what IS there"


def test_a_missing_sheet_does_not_report_every_field_as_missing(tmp_path):
    """The old failure mode: 'missing required fields (company_name, ticker,
    …)', which is the same message as a dozen other causes."""
    src = _write(tmp_path / "dennis_data.xlsx", GOOD_ROWS, sheet="Sheet1")
    text = check_export(src).render()
    assert "company_name" not in text and "market_cap" not in text


# --------------------------------------------------------------------------
# Formulas that never resolved
# --------------------------------------------------------------------------


def test_a_workbook_of_live_formulas_is_diagnosed(tmp_path):
    """Under data_only=True an uncalculated formula reads as None, so this
    would otherwise look like an empty workbook rather than the wrong export."""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Snapshot"
    ws.append(["", "field_key", "", "Value (auto)", "", "Priority"])
    for key in ("company_name", "ticker", "price"):
        ws.append(["", key, "", f'=CIQ($C$3,"IQ_{key.upper()}")', "", "Required"])
    src = tmp_path / "dennis_data.xlsx"
    wb.save(src)

    check = check_export(src)
    assert not check.ok
    problem = check.blocking[0]
    assert problem.kind == "formulas_only"
    assert "Values" in problem.message, "name the Paste Special fix"


def test_an_all_blank_workbook_is_distinguished_from_a_formula_one(tmp_path):
    src = _write(tmp_path / "dennis_data.xlsx",
                 [("company_name", ""), ("ticker", ""), ("price", "")])
    check = check_export(src)
    assert not check.ok
    assert check.blocking[0].kind == "no_fields"
    assert "before the add-in finished" in check.blocking[0].message


def test_a_workbook_that_is_not_a_workbook(tmp_path):
    src = tmp_path / "dennis_data.xlsx"
    src.write_bytes(b"this is not a zip container")
    check = check_export(src)
    assert not check.ok
    assert check.blocking[0].kind == "unreadable"


# --------------------------------------------------------------------------
# Wrong ticker
# --------------------------------------------------------------------------


def test_a_workbook_for_the_wrong_company_is_refused(tmp_path):
    src = _write(tmp_path / "dennis_data.xlsx", GOOD_ROWS)     # EXMPL
    check = check_export(src, expect_ticker="PLTR")
    assert not check.ok
    problem = check.blocking[0]
    assert problem.kind == "wrong_ticker"
    assert "EXMPL" in problem.message and "PLTR" in problem.message
    assert "/short EXMPL" in problem.message, "offer the way out"


def test_the_ticker_check_ignores_case_and_padding(tmp_path):
    src = _write(tmp_path / "dennis_data.xlsx",
                 [("ticker", "  exmpl "), *GOOD_ROWS[2:]])
    assert check_export(src, expect_ticker="EXMPL").ok


def test_a_workbook_with_no_ticker_only_warns(tmp_path):
    src = _write(tmp_path / "dennis_data.xlsx", GOOD_ROWS[:1] + GOOD_ROWS[2:])
    check = check_export(src, expect_ticker="EXMPL")
    assert check.ok
    assert any(p.kind == "wrong_ticker" for p in check.warnings)


# --------------------------------------------------------------------------
# Staleness — from the workbook's own as-of date
# --------------------------------------------------------------------------


def test_a_stale_workbook_is_flagged_against_its_own_as_of_date(tmp_path):
    src = _write(tmp_path / "dennis_data.xlsx", GOOD_ROWS)   # as of 2026-07-20
    check = check_export(src, max_age_days=10, today=dt.date(2026, 8, 30))
    assert check.ok, "stale is a warning, not a refusal"
    stale = [p for p in check.warnings if p.kind == "stale"]
    assert stale and "41 days old" in stale[0].message
    assert "2026-07-20" in stale[0].message


def test_a_fresh_workbook_says_nothing(tmp_path):
    src = _write(tmp_path / "dennis_data.xlsx", GOOD_ROWS)
    check = check_export(src, max_age_days=10, today=dt.date(2026, 7, 22))
    assert check.problems == []


def test_a_workbook_with_no_as_of_date_is_noted(tmp_path):
    src = _write(tmp_path / "dennis_data.xlsx",
                 [r for r in GOOD_ROWS if r[0] != "as_of_date"])
    check = check_export(src, max_age_days=10)
    assert check.ok
    assert any("no as-of date" in p.message for p in check.warnings)


@pytest.mark.parametrize("written,expected_age", [
    ("2026-07-20", 41),
    ("20/07/2026", 41),
    ("07/20/2026", 41),
])
def test_the_as_of_date_is_read_in_the_shapes_excel_writes_it(
        tmp_path, written, expected_age):
    src = _write(tmp_path / f"{expected_age}.xlsx",
                 [("as_of_date", written), *GOOD_ROWS[3:]])
    check = check_export(src, max_age_days=10, today=dt.date(2026, 8, 30))
    stale = [p for p in check.warnings if p.kind == "stale"]
    assert stale and f"{expected_age} days old" in stale[0].message


# --------------------------------------------------------------------------
# The bot end: a bad upload must not destroy a good workbook
# --------------------------------------------------------------------------


def _core(settings):
    from bot.handlers import BotCore

    return BotCore(settings)


def test_a_good_upload_lands_and_hands_back_the_prompts(settings, tmp_path):
    core = _core(settings)
    core.start_lane(1, "short", "EXMPL")
    reply = core.handle_upload(1, "dennis_data.xlsx", FIXTURE.read_bytes())

    assert "saved dennis_data.xlsx" in reply.text
    ws = core.context.get(1)
    assert (ws.path / "dennis_data.xlsx").exists()
    assert any(f.name.startswith("prompt_") for f in reply.files), reply.files


def test_a_wrong_ticker_upload_leaves_the_existing_workbook_alone(settings, tmp_path):
    """The property that matters most: the operator uploaded the wrong file,
    and the right one is still there."""
    core = _core(settings)
    core.start_lane(2, "short", "EXMPL")
    ws = core.context.get(2)
    core.handle_upload(2, "dennis_data.xlsx", FIXTURE.read_bytes())
    good = (ws.path / "dennis_data.xlsx").read_bytes()

    wrong = _write(tmp_path / "other.xlsx",
                   [("company_name", "Palantir"), ("ticker", "PLTR"),
                    ("as_of_date", "2026-07-20"), ("price", 158.4),
                    ("market_cap", 1.0), ("shares_out", 1.0)])
    reply = core.handle_upload(2, "dennis_data.xlsx", wrong.read_bytes())

    assert "NOT updated" in reply.text
    assert "PLTR" in reply.text and "EXMPL" in reply.text
    assert (ws.path / "dennis_data.xlsx").read_bytes() == good, \
        "the good workbook was overwritten by a rejected upload"


def test_an_unresolved_upload_is_refused_with_the_reason(settings, tmp_path):
    core = _core(settings)
    core.start_lane(3, "short", "EXMPL")
    bad = _write(tmp_path / "bad.xlsx",
                 [("company_name", "#CIQINACTIVE"), ("ticker", "EXMPL"),
                  ("as_of_date", "2026-07-20"), ("price", 158.4),
                  ("market_cap", 1.0), ("shares_out", 1.0)])

    reply = core.handle_upload(3, "dennis_data.xlsx", bad.read_bytes())
    assert "NOT updated" in reply.text
    assert "#CIQINACTIVE" in reply.text
    assert not (core.context.get(3).path / "dennis_data.xlsx").exists()


def test_a_rejected_upload_leaves_no_scratch_files(settings, tmp_path):
    core = _core(settings)
    core.start_lane(4, "short", "EXMPL")
    ws = core.context.get(4)
    junk = tmp_path / "junk.xlsx"
    junk.write_bytes(b"not a workbook")

    core.handle_upload(4, "dennis_data.xlsx", junk.read_bytes())
    leftovers = [p.name for p in ws.path.iterdir() if p.name.startswith(".upload_")]
    assert leftovers == [], leftovers


def test_the_rejection_message_names_the_operators_filename(settings, tmp_path):
    """The staging file is an implementation detail; the operator uploaded
    `dennis_data.xlsx` and that is what the message should talk about."""
    core = _core(settings)
    core.start_lane(5, "short", "EXMPL")
    bad = _write(tmp_path / "bad.xlsx", GOOD_ROWS[:1] + [("ticker", "PLTR")] + GOOD_ROWS[2:])

    reply = core.handle_upload(5, "dennis_data.xlsx", bad.read_bytes())
    assert ".upload_" not in reply.text
    assert "dennis_data.xlsx" in reply.text


def test_a_stale_upload_still_lands_but_says_so(settings, tmp_path):
    """Stale is a warning: the operator may be deliberately covering an old
    print, and refusing the file would be the tool overruling them."""
    core = _core(settings)
    core.start_lane(6, "short", "EXMPL")
    old = _write(tmp_path / "old.xlsx",
                 [("company_name", "Example Industries"), ("ticker", "EXMPL"),
                  ("as_of_date", "2020-01-01"), ("price", 158.4),
                  ("market_cap", 1.0), ("shares_out", 1.0)])

    reply = core.handle_upload(6, "dennis_data.xlsx", old.read_bytes())
    assert (core.context.get(6).path / "dennis_data.xlsx").exists()
    assert "days old" in reply.text
