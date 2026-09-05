"""The valuation pack: region slots, the `{t, median}` pair, and the routing.

`tables/multiples-strip` is the first plate in the library whose slots do not
all take a string. `marker-N` is a REGION carrying a pair of numbers, and every
test here exists because one specific way of getting that wrong renders a frame
that looks fine and says something false:

* a string typeset into a region — a plate cannot know a percentile, so a
  director who writes one has invented the position rather than read it;
* a `t` clamped "to be safe" — the ends of the rail are the peer set's 10th and
  90th percentile, so off the range is the most quotable row on the plate and
  clamping deletes the finding;
* the portrait strip treated as the landscape one with rows removed — it has
  one fewer COLUMN too, and the peer number reaches it only through the
  marker's `median`.
"""

from __future__ import annotations

import pytest

from pipeline.plates import load_plates

STRIP_16 = "tables/multiples-strip-16x9"
STRIP_9 = "tables/multiples-strip-9x16"
BRIDGE = "structure/multiple-bridge-16x9"


@pytest.fixture()
def reg(settings):
    return load_plates(settings.assets_dir)


# --------------------------------------------------------------------------
# §0 — one normalised frame shape across the whole library.
# --------------------------------------------------------------------------
def test_every_family_loads_to_one_normalised_frame_shape(reg):
    """All fourteen families, one shape out — asserted on the LOADED plate.

    Deliberately not asserted against raw JSON. The delivery has twice shipped
    a family whose `frames[]` entries differed from the rest of the library in
    which optional keys were present and whether an absent one was `null` or
    `""`, and both times every downstream reader had to cope. The contract this
    codebase actually depends on is what comes out of the registry: a `boil`
    that is always an int, a `tag` that is always a string, and a
    `base_is_frame` that is `""` or a real tag and never `None`.
    """
    assert len(reg.families()) == 14, reg.families()
    for key, plate in reg.assets.items():
        assert plate.frames, f"{key} declares no frames"
        assert isinstance(plate.base_is_frame, str), key
        for frame in plate.frames:
            assert isinstance(frame.boil, int), f"{key} {frame.png}: boil"
            assert isinstance(frame.tag, str), f"{key} {frame.png}: tag"
            assert isinstance(frame.png, str) and frame.png, f"{key}: png"
            assert isinstance(frame.mouth_open, bool), key
            assert isinstance(frame.bob, int), key
        assert plate.frame_count == len(plate.frames), key


def test_an_unknown_manifest_header_key_does_not_break_ingest(reg):
    """Family headers are prose ABOUT the family and gain keys between packs.

    09b added `maxCharsNote` to all fourteen and `budgetNote` to the strip. A
    reader with a strict header schema rejects a delivery for documenting
    itself, so nothing here reads the header at all — the assets are the
    contract. This asserts the new keys are present AND that the kit still
    loaded, which is the whole claim.
    """
    import json
    from pathlib import Path

    kit = Path(__file__).resolve().parents[1] / "kit"
    for family in ("tables", "structure"):
        header = json.loads((kit / family / "manifest.json")
                            .read_text(encoding="utf-8"))
        assert "maxCharsNote" in header, family
        assert header.get("assets"), family
    assert len(reg.assets) > 0


def test_the_three_new_keys_are_in_the_kit(reg):
    for key in (STRIP_16, STRIP_9, BRIDGE):
        assert key in reg, f"{key} did not survive ingest"
    assert reg.require(STRIP_16).rows == 6
    assert reg.require(STRIP_9).rows == 3
    # 16:9 carries the median column; the portrait re-author does not.
    assert reg.require(STRIP_16).slot("median-1") is not None
    assert reg.require(STRIP_9).slot("median-1") is None
    assert reg.require(STRIP_9).slot("head-median") is None


# --------------------------------------------------------------------------
# §6.6 — a string bound to a region slot raises.
# --------------------------------------------------------------------------
def test_a_string_bound_to_a_region_slot_is_rejected(reg):
    from pipeline.plate_tags import build_fill

    fill = build_fill(
        reg, f"{STRIP_16} | label-1=P/E | subject-1=58.2x "
             f"| marker-1=82nd percentile", aspect="16x9")
    assert not fill.ok
    assert any("region" in p and "marker-1" in p for p in fill.problems), \
        fill.problems


def test_a_region_slot_is_never_typeset_even_with_commas_in_it(reg):
    """The comma is not a licence.

    A region that declares no renderer takes nothing at all, and `figure=he,
    seated` is prose rather than a series — which the series branch would
    otherwise have accepted as two figures.
    """
    from pipeline.plate_tags import build_fill

    room = next((k for k in reg.keys()
                 if reg.assets[k].slot("host-anchor") is not None), None)
    assert room, "no plate in the kit reserves a host anchor"
    fill = build_fill(reg, f"{room} | host-anchor=he, seated")
    assert not fill.ok
    assert any("host-anchor" in p for p in fill.problems), fill.problems


def test_a_numeric_pair_bound_to_a_text_slot_is_rejected(reg):
    from pipeline.plate_tags import build_fill

    fill = build_fill(reg, f"{STRIP_16} | label-1=P/E "
                           f"| subject-1=t:0.82,median:0.41", aspect="16x9")
    assert not fill.ok
    assert any("subject-1" in p and "range mark" in p for p in fill.problems), \
        fill.problems


# --------------------------------------------------------------------------
# §6.6 — the pair itself.
# --------------------------------------------------------------------------
def test_a_marker_without_median_is_rejected(reg):
    from pipeline.plate_tags import build_fill

    fill = build_fill(reg, f"{STRIP_9} | label-1=P/E | subject-1=58.2x "
                           f"| marker-1=t:0.82", aspect="9x16")
    assert not fill.ok
    assert any("median" in p for p in fill.problems), fill.problems


def test_a_marker_without_t_is_rejected(reg):
    from pipeline.plate_tags import build_fill

    fill = build_fill(reg, f"{STRIP_16} | label-1=P/E | subject-1=58.2x "
                           f"| marker-1=median:0.41", aspect="16x9")
    assert not fill.ok
    assert any("no t" in p for p in fill.problems), fill.problems


def test_t_outside_0_1_survives_the_parser_unclamped(reg):
    """`t = 1.4` is a reading, and the parser is not allowed to round it off.

    The rail ends are p10 and p90 rather than min and max, so a subject priced
    above every peer genuinely lands past 1. Clamping here would draw a
    plausible row that says something weaker than the truth.
    """
    from pipeline.plate_tags import build_fill, parse_marker

    fill = build_fill(reg, f"{STRIP_16} | label-1=P/S | subject-1=141.6x "
                           f"| marker-1=t:1.4,median:0.44", aspect="16x9")
    assert fill.ok, fill.problems
    pair, why = parse_marker(fill.values["marker-1"])
    assert why == ""
    assert pair.t == 1.4, "the parser clamped a real reading"
    assert pair.off_range
    assert any("outside the peer range" in w for w in fill.warnings), fill.warnings

    below, _ = parse_marker("t:-0.3,median:0.5")
    assert below.t == -0.3


# --------------------------------------------------------------------------
# §6.6 — `t = 1.4` RENDERS: dot on the end tick, chevron past it.
# --------------------------------------------------------------------------
def test_an_off_range_reading_renders_inside_its_region(settings, reg):
    """It draws, it draws MORE than an in-range row (the chevron), and none of
    it lands outside the box the plate reserved."""
    import numpy as np
    from pipeline.chart import draw_declared
    from pipeline.plate_frames import render_still
    from pipeline.plate_tags import build_fill

    fill = build_fill(
        reg, f"{STRIP_16} | label-1=P/E | subject-1=58.2x | median-1=24.1x "
             f"| marker-1=t:0.5,median:0.41 "
             f"| label-2=P/S | subject-2=141.6x | median-2=6.1x "
             f"| marker-2=t:1.4,median:0.44", aspect="16x9")
    assert fill.ok, fill.problems

    plate = reg.require(STRIP_16)
    before = render_still(plate, fill.values, settings, reg).convert("RGBA")
    after = before.copy()
    assert draw_declared(reg, plate, fill.values, after, seed=plate.key)

    changed = (abs(np.array(after).astype(int)
                   - np.array(before).astype(int)).sum(-1) > 12)

    def ink(name):
        x, y, w, h = plate.slot(name).scaled()
        return int(changed[y:y + h, x:x + w].sum())

    assert ink("marker-1") > 0, "the in-range row drew no mark"
    assert ink("marker-2") > ink("marker-1"), \
        "the off-range row drew no more ink than the in-range one — the " \
        "chevron is missing, and the row reads as if it sat on the end"

    # NOTHING PAINTS OUTSIDE THE RESERVED REGIONS. The mark is inset by its own
    # radius precisely so t = 1 sits tangent to the end tick instead of half
    # outside the box.
    inside = np.zeros(changed.shape, bool)
    for i in range(1, plate.rows + 1):
        x, y, w, h = plate.slot(f"marker-{i}").scaled()
        inside[y:y + h, x:x + w] = True
    assert int((changed & ~inside).sum()) == 0, \
        "the range mark painted outside the region the plate reserved"


def test_every_rail_the_director_filled_gets_its_own_mark(settings, reg):
    """Six rails, six pairs, six marks — not one mark and five empty rails."""
    import numpy as np
    from pipeline.chart import draw_declared
    from pipeline.plate_frames import render_still
    from pipeline.plate_tags import build_fill

    parts = [STRIP_16]
    for i in range(1, 7):
        parts.append(f"label-{i}=M{i} | subject-{i}={i}.0x "
                     f"| marker-{i}=t:0.{i}5,median:0.4")
    fill = build_fill(reg, " | ".join(parts), aspect="16x9")
    assert fill.ok, fill.problems

    plate = reg.require(STRIP_16)
    before = render_still(plate, fill.values, settings, reg).convert("RGBA")
    after = before.copy()
    draw_declared(reg, plate, fill.values, after, seed=plate.key)
    changed = (abs(np.array(after).astype(int)
                   - np.array(before).astype(int)).sum(-1) > 12)
    for i in range(1, 7):
        x, y, w, h = plate.slot(f"marker-{i}").scaled()
        assert int(changed[y:y + h, x:x + w].sum()) > 0, f"marker-{i} drew nothing"


def test_an_unwritten_rail_draws_nothing(settings, reg):
    """A metric with no peer data keeps its rail and gets no mark.

    An empty cell in this library means NO DATA. A mark at a default position
    would be the renderer inventing a number.
    """
    import numpy as np
    from pipeline.chart import draw_declared
    from pipeline.plate_frames import render_still
    from pipeline.plate_tags import build_fill

    fill = build_fill(reg, f"{STRIP_16} | label-1=P/E | subject-1=58.2x",
                      aspect="16x9")
    assert fill.ok, fill.problems
    plate = reg.require(STRIP_16)
    before = render_still(plate, fill.values, settings, reg).convert("RGBA")
    after = before.copy()
    assert not draw_declared(reg, plate, fill.values, after, seed=plate.key)
    assert not (abs(np.array(after).astype(int)
                    - np.array(before).astype(int)).sum(-1) > 12).any()


# --------------------------------------------------------------------------
# §6.6 — routing: capacity, the dropped column, and the missing aspect.
# --------------------------------------------------------------------------
def test_a_six_row_strip_in_a_short_is_rejected(reg):
    from pipeline.plate_tags import build_fill

    fill = build_fill(reg, f"{STRIP_9} | label-1=P/E | subject-1=58x "
                           f"| label-6=FCF yield | subject-6=1.1%",
                      aspect="9x16")
    assert not fill.ok
    assert any("3 rows" in p for p in fill.problems), fill.problems


def test_median_on_a_portrait_strip_is_rejected(reg):
    """And the message says where the peer number goes instead."""
    from pipeline.plate_tags import build_fill

    for written in ("median-3=24.1x", "head-median=Peer median"):
        fill = build_fill(reg, f"{STRIP_9} | label-1=P/E | subject-1=58x "
                               f"| {written}", aspect="9x16")
        assert not fill.ok, written
        assert any("marker-N" in p for p in fill.problems), fill.problems


def test_a_dropped_column_and_a_row_past_capacity_say_different_things(reg):
    """`median-7` on the LANDSCAPE strip is over capacity, not a missing column.

    The two failures look identical from the slot table — neither name is a
    declared slot — and telling a director the 16:9 plate "carries no median
    column at all" sends them to fix the wrong thing.
    """
    from pipeline.plate_tags import build_fill

    over = build_fill(reg, f"{STRIP_16} | label-1=P/E | subject-1=58x "
                           f"| median-7=24.1x", aspect="16x9")
    assert not over.ok
    assert any("6 rows" in p for p in over.problems), over.problems
    assert not any("no median column" in p for p in over.problems), over.problems

    dropped = build_fill(reg, f"{STRIP_9} | label-1=P/E | subject-1=58x "
                              f"| median-1=24.1x", aspect="9x16")
    assert not dropped.ok
    assert any("no median column" in p for p in dropped.problems), dropped.problems


def test_the_bridge_is_refused_in_a_portrait_cut(reg):
    from pipeline.plate_tags import build_fill, check_bound

    fill = build_fill(reg, f"{BRIDGE} | kicker=THE DENOMINATOR "
                           f"| step-1-figure=58.2x | step-1-label=Trailing",
                      aspect="9x16")
    assert not fill.ok
    assert any("9x16" in p for p in fill.problems), fill.problems

    # And again at the gate, where the values arrive already bound.
    bound = check_bound(reg, BRIDGE, {"kicker": "THE DENOMINATOR"},
                        aspect="9x16")
    assert not bound.ok
    assert any("9x16" in p for p in bound.problems), bound.problems


def test_the_bridge_is_allowed_in_a_landscape_cut(reg):
    from pipeline.plate_tags import build_fill

    fill = build_fill(
        reg, f"{BRIDGE} | kicker=WHAT COMES OUT | step-1-figure=58.2x "
             f"| step-1-label=Trailing P/E | step-2-figure=44.1x "
             f"| step-2-label=Adjusted | step-3-figure=31.5x "
             f"| step-3-label=Forward P/E | link-1-note=one-off charge "
             f"| link-2-note=consensus growth | caption=The denominator moved.",
        aspect="16x9")
    assert fill.ok, fill.problems


def test_the_valuation_chapter_may_reach_for_all_three(reg):
    for key in (STRIP_16, STRIP_9, BRIDGE):
        assert reg.chapter_allows("valuation", key), key


def test_the_strip_and_the_peer_strip_are_not_variants(reg):
    """Rows are metrics on one and companies on the other.

    They are inverses, and the aspect resolver must never treat one as the
    other's missing half.
    """
    assert reg.aspect_key("tables/multiples-strip", "9x16") == STRIP_9
    assert reg.aspect_key("structure/multiple-bridge", "9x16") is None
    assert reg.aspect_key("structure/multiple-bridge", "16x9") == BRIDGE


# --------------------------------------------------------------------------
# §6.6 — Snapshot!D50/D51 parse as fractions and tolerate blank.
# --------------------------------------------------------------------------
def test_ownership_parses_as_a_fraction(fixtures_dir, tmp_path):
    import shutil

    from pipeline.company_data import load_company_data

    shutil.copy(fixtures_dir / "company_data" / "dennis_data.xlsx",
                tmp_path / "dennis_data.xlsx")
    data = load_company_data(tmp_path)
    for field in ("insider_own", "institutional_own"):
        value = data.values.get(field)
        assert isinstance(value, float), f"{field} is {value!r}"
        assert 0.0 <= value <= 1.0, \
            f"{field} = {value} is not a fraction — D50/D51 are 0-1, the same " \
            f"convention D49 short_interest uses, and dividing again downstream " \
            f"is a 100x error in a sentence that reads perfectly well"


def test_blank_ownership_is_missing_and_never_zero(fixtures_dir, tmp_path):
    """The vendor fields do not resolve on every licence, and D blanks anything
    that lands outside 0-1 rather than shipping a nonsense number.

    Missing must stay missing all the way to the writer: "0% insider ownership"
    is a claim about the company, and the data never made it.
    """
    import shutil

    import openpyxl

    from pipeline.company_data import load_company_data

    src = tmp_path / "dennis_data.xlsx"
    shutil.copy(fixtures_dir / "company_data" / "dennis_data.xlsx", src)
    book = openpyxl.load_workbook(src)
    sheet = book["Snapshot"]
    blanked = 0
    for row in range(1, sheet.max_row + 1):
        if sheet.cell(row, 2).value in ("insider_own", "institutional_own"):
            sheet.cell(row, 4).value = None
            blanked += 1
    assert blanked == 2, "the fixture no longer carries both ownership rows"
    book.save(src)

    data = load_company_data(src.parent)
    for field in ("insider_own", "institutional_own"):
        assert data.values.get(field) in (None, ""), data.values.get(field)
        assert field in data.missing
    # And it is absent from what the writer is given, rather than present as 0.
    block = data.as_prompt_block()
    assert "insider_own" not in block
    assert "institutional_own" not in block


def test_the_rail_columns_are_read_and_never_clamped(fixtures_dir, tmp_path):
    """`Peers!G:J` — and `t` is NOT column D.

    Column D is a rank. Feeding it to the rail computes every median tick to
    0.5 and puts them all dead centre, which is the one comparison the plate
    exists to make.
    """
    import shutil

    from pipeline.company_data import load_company_data

    shutil.copy(fixtures_dir / "company_data" / "dennis_data.xlsx",
                tmp_path / "dennis_data.xlsx")
    rows = load_company_data(tmp_path).peer_percentiles
    assert rows, "no peer percentile block in the fixture"
    for row in rows:
        for column in ("peer_low", "peer_high", "t", "t_median"):
            assert column in row, f"{row['metric']} has no {column}"

    metrics = {r["metric"] for r in rows}
    assert "Forward P/E" in metrics and "Forward PEG" in metrics, \
        "rows 68-69 carry the two forward multiples move 2 reaches for"

    by_metric = {r["metric"]: r for r in rows}
    assert by_metric["P/S (TTM)"]["t"] > 1.0, \
        "the off-range row was clamped somewhere between the sheet and here"
    # t is a position on a value axis, and the rank is a different number.
    assert by_metric["P/S (TTM)"]["t"] != by_metric["P/S (TTM)"]["percentile"]


# --------------------------------------------------------------------------
# §6.5 — the gate fails a chapter that skips move 3.
# --------------------------------------------------------------------------
def test_the_gate_fails_a_chapter_that_skips_the_peer_comparison(settings):
    from pipeline.gates import valuation_moves

    class _Script:
        narration = ("It trades on 26 times forward earnings. At that price "
                     "the market is priced for 25% growth a year, forever, "
                     "and the business has delivered nine.")
        events: list = []

    findings = valuation_moves(_Script(), settings)
    assert findings, "forward multiples into a reverse DCF with no peer move"
    assert findings[0].severity == "block"
    assert "move 3" in findings[0].message


def test_the_gate_is_satisfied_by_the_strip_alone(settings):
    """The plate IS the move. A director who reached for it made the
    comparison whether or not the narration used the word "peer"."""
    from pipeline.gates import valuation_moves
    from pipeline.models import TagEvent, TagType

    class _Script:
        narration = ("It trades on 26 times forward earnings. The price is "
                     "priced for 25% growth a year, forever.")
        events = [TagEvent(type=TagType.PLATE, payload=STRIP_16,
                           char_offset=0, raw_offset=0)]

    assert valuation_moves(_Script(), settings) == []


def test_the_gate_leaves_a_chapter_that_never_gets_there_alone(settings):
    """A script that reaches neither move 2 nor move 4 is a different shape,
    and is not failed here for being short."""
    from pipeline.gates import valuation_moves

    class _Script:
        narration = "Revenue grew nine percent and the margin held."
        events: list = []

    assert valuation_moves(_Script(), settings) == []
