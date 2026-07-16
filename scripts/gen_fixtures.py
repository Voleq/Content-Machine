"""Regenerate binary/derived fixtures committed to the repo.

Run from the repo root:  .venv/bin/python scripts/gen_fixtures.py

Produces:
  fixtures/company_data/dennis_data.xlsx   (the v3 export, filled for EXMPL:
                                            Snapshot · History · Dashboard ·
                                            Valuation [+ WACC/reverse-DCF] ·
                                            Peers [+ self-scoring percentiles] ·
                                            News)
  fixtures/company_data/dennis_data.csv    (CSV flavour — snapshot only)
  fixtures/tts/alignment_sample.json       (ElevenLabs-style char alignment)
  fixtures/prices/EXMPL.json               (deterministic price history)

The canonical template (templates/dennis_data_template.xlsx) is the real
v3 workbook the operator refreshes — it is committed as-is, not built here.
"""

from __future__ import annotations

import datetime as _dt
import json
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

NA = "#N/A"  # coerces to None on read

# --- Snapshot: (section, field_key, label, value, mnemonic-hint) --------------
EXMPL_SNAPSHOT = [
    ("Identity", "company_name", "Company Name", "Example Corp", "TR.CommonName"),
    ("", "ticker", "Ticker / RIC", "EXMPL", ""),
    ("", "exchange", "Exchange", "NASDAQ Global Select Market", "TR.ExchangeName"),
    ("", "sector", "GICS Sector", "Information Technology", "TR.GICSSector"),
    ("", "industry", "Industry", "Application Software", "TR.GICSIndustry"),
    ("", "country", "Country", "United States of America", "TR.HeadquartersCountry"),
    ("", "currency", "Reporting Currency", "USD", "TR.FinancialCurrency"),
    ("", "as_of_date", "As-of Date", "2026-07-01", ""),
    ("Price & Size", "price", "Price", 84.20, "TR.PriceClose"),
    ("", "market_cap", "Market Cap", 30_700_000_000, "TR.CompanyMarketCap"),
    ("", "enterprise_value", "Enterprise Value", 31_900_000_000, "TR.EV"),
    ("", "shares_out", "Shares Outstanding", 364_600_000, "TR.SharesOutstanding"),
    ("", "avg_volume_3m", "Avg Daily Volume (3m)", 8_000_000, "TR.AvgDailyVolume3M"),
    ("", "beta", "Beta (5y)", 1.6, "TR.Beta"),
    ("", "week52_high", "52-Week High", 120.0, "TR.Price52WeekHigh"),
    ("", "week52_low", "52-Week Low", 40.0, "TR.Price52WeekLow"),
    ("", "pct_from_52w_high", "% From 52w High", -29.83, "computed"),
    ("Valuation (now)", "pe_ttm", "P/E (TTM)", NA, "TR.PE"),
    ("", "forward_pe", "Forward P/E", NA, "TR.FwdPE"),
    ("", "ps_ttm", "P/S (TTM)", 62.0, "TR.PriceToSalesPerShare"),
    ("", "ev_ebitda", "EV / EBITDA", NA, "TR.EVToEBITDA"),
    ("", "ev_sales", "EV / Sales", 64.3, "TR.EVToSales"),
    ("", "ev_fcf", "EV / FCF", NA, "TR.EVToFCF"),
    ("", "pb", "P / B", 18.0, "TR.PriceToBVPerShare"),
    ("", "p_fcf", "P / FCF", NA, "TR.PriceToFCFPerShare"),
    ("", "peg", "PEG", NA, "TR.PEG"),
    ("", "earnings_yield", "Earnings Yield (%)", NA, "computed"),
    ("", "fcf_yield", "FCF Yield (%)", -5.0, "TR.FCFYield"),
    ("", "dividend_yield", "Dividend Yield (%)", 0.0, "TR.DividendYield"),
    ("", "buyback_yield", "Buyback Yield (%)", 0.0, "TR.BuybackYield"),
    ("", "shareholder_yield", "Shareholder Yield (%)", 0.0, "computed"),
    ("Balance (now)", "cash_st", "Cash & ST Investments", 410_000_000, "TR.CashAndSTInvestments"),
    ("", "total_debt_now", "Total Debt", 1_450_000_000, "TR.TotalDebtOutstanding"),
    ("", "net_debt_now", "Net Debt", 1_040_000_000, "computed"),
    ("", "net_debt_ebitda_now", "Net Debt / EBITDA", 9.8, "TR.NetDebtToEBITDA"),
    ("", "debt_to_equity", "Debt / Equity (%)", 140.0, "TR.TotalDebtToTotalEquityPct"),
    ("", "current_ratio", "Current Ratio", 1.4, "TR.CurrentRatio"),
    ("", "quick_ratio", "Quick Ratio", 1.1, "TR.QuickRatio"),
    ("", "interest_coverage", "Interest Coverage", 0.9, "TR.InterestCoverageRatio"),
    ("", "goodwill_intang", "Goodwill & Intangibles", NA, "TR.GoodwillGross"),
    ("", "tangible_bv", "Tangible BV / Share", NA, "TR.TangibleBVPerShare"),
    ("Ownership & Risk", "short_interest", "Short Interest (% float)", 11.0, "TR.ShortInterestPercentOfFloat"),
    ("", "insider_own", "Insider Ownership (%)", 8.0, "TR.InsiderPercentHeld"),
    ("", "institutional_own", "Institutional Own (%)", 62.0, "TR.InstitutionsPercentHeld"),
]

# --- History: 6 periods, oldest -> newest (FY-4..FY-0, LTM) -------------------
EXMPL_PERIODS = ["FY-4", "FY-3", "FY-2", "FY-1", "FY-0", "LTM"]
M = 1_000_000
EXMPL_HISTORY = [
    ("revenue", "Revenue", [400*M, 452*M, 471*M, 491*M, 496*M, 496*M]),
    ("gross_profit", "Gross Profit", [252*M, 280*M, 287*M, 290*M, 288*M, 288*M]),
    ("gross_margin", "Gross Margin %", [63.0, 62.0, 61.0, 59.0, 58.0, 58.0]),
    ("operating_income", "Operating Income", [-8*M, -23*M, -38*M, -49*M, -60*M, -60*M]),
    ("operating_margin", "Operating Margin %", [-2.0, -5.0, -8.0, -10.0, -12.0, -12.0]),
    ("ebitda", "EBITDA", [10*M, -5*M, -20*M, -35*M, -45*M, -45*M]),
    ("net_income", "Net Income", [-8*M, -25*M, -49*M, -70*M, -89*M, -89*M]),
    ("net_margin", "Net Margin %", [-2.0, -5.5, -10.4, -14.3, -18.0, -18.0]),
    ("operating_cf", "Operating Cash Flow", [15*M, 2*M, -3*M, -8*M, -9*M, -9*M]),
    ("capex", "CapEx", [-3*M, -4*M, -3*M, -3*M, -6*M, -6*M]),
    ("fcf", "Free Cash Flow", [12*M, -2*M, -6*M, -11*M, -15*M, -15*M]),
    ("fcf_margin", "FCF Margin %", [3.0, -0.4, -1.3, -2.2, -3.0, -3.0]),
    ("sbc", "Stock-Based Comp", [40*M, 60*M, 80*M, 95*M, 110*M, 110*M]),
    ("cash", "Cash & Equivalents", [700*M, 610*M, 520*M, 465*M, 410*M, 410*M]),
    ("total_debt", "Total Debt", [900*M, 1050*M, 1200*M, 1350*M, 1450*M, 1450*M]),
    ("net_debt", "Net Debt", [200*M, 440*M, 680*M, 885*M, 1040*M, 1040*M]),
    ("total_equity", "Total Equity", [1200*M, 1150*M, 1080*M, 1030*M, 1036*M, 1036*M]),
    ("roic", "ROIC %", [-0.4, -1.1, -2.1, -2.9, -3.6, -3.6]),
    ("diluted_shares", "Diluted Shares", [298*M, 315*M, 330*M, 346*M, 364.6*M, 364.6*M]),
    ("shares_yoy", "Share Count YoY %", [NA, 5.7, 4.8, 4.8, 5.4, 0.0]),
]

# --- Dashboard: label -> value (+ flags) --------------------------------------
EXMPL_DASHBOARD = [
    ("Company", "Example Corp"),
    ("Price", 84.20),
    ("Market cap", 30_700_000_000),
    ("% from 52w high", -29.83),
    ("Revenue (LTM)", 496*M),
    ("Revenue 4y CAGR", 5.5),
    ("Gross margin (LTM)", 58.0),
    ("Operating margin (LTM)", -12.0),
    ("Net margin (LTM)", -18.0),
    ("FCF (LTM)", -15*M),
    ("FCF margin (LTM)", -3.0),
    ("FCF yield", -5.0),
    ("Net debt / EBITDA", 9.8),
    ("ROIC (LTM)", -3.6),
    ("Share count 4y CAGR", 5.2),
    ("P/E (TTM)", NA),
    ("EV/EBITDA", NA),
    ("Quick flags", ""),
    ("Profitable? (LTM NI>0)", "no"),
    ("FCF positive? (LTM)", "no"),
    ("Growing? (rev CAGR>0)", "yes"),
    ("Diluting? (share CAGR>1%)", "yes"),
    ("Over-levered? (ND/EBITDA>3x)", "yes"),
]

# --- Valuation: inputs + bear/base/bull (P/S based, EPS is negative) ----------
_REV_PS = 496 * M / 364.6e6  # ~1.36 revenue/share
EXMPL_VAL_INPUTS = [("Current price", 84.20), ("LTM EPS", NA),
                    ("LTM FCF / share", -0.041)]
EXMPL_VAL_SCENARIOS = [
    # scenario, metric, (basis), exit multiple, implied price, upside %
    ("Bear", "P/S", "Sales/sh", 3.0, round(3.0 * _REV_PS, 2), round((3.0*_REV_PS/84.20-1)*100, 1)),
    ("Base", "P/S", "Sales/sh", 6.0, round(6.0 * _REV_PS, 2), round((6.0*_REV_PS/84.20-1)*100, 1)),
    ("Bull", "P/S", "Sales/sh", 12.0, round(12.0 * _REV_PS, 2), round((12.0*_REV_PS/84.20-1)*100, 1)),
]

# --- Peers: EXMPL's logistics-software comps ----------------------------------
EXMPL_PEERS_HEADER = ["Peer (auto)", "Price", "Market Cap", "P/E", "EV/EBITDA",
                      "P/S", "Rev Growth %", "Gross Mgn %", "Net Mgn %",
                      "FCF Yield %", "NetDebt/EBITDA"]
EXMPL_PEERS = [
    ("Freightwave Systems Inc", 142.0, 12_400_000_000, 34.0, 22.0, 9.1, 18.0, 71.0, 6.0, 2.1, 0.4),
    ("DepotOps Corp", 63.0, 4_100_000_000, NA, 41.0, 5.4, 9.0, 62.0, -4.0, 0.6, 1.8),
    ("RouteLogic Holdings", 208.0, 28_900_000_000, 41.0, 26.0, 11.2, 22.0, 76.0, 12.0, 3.0, 0.1),
    ("CargoCloud Inc", 19.0, 1_800_000_000, NA, NA, 3.2, 6.0, 58.0, -22.0, NA, 2.6),
]

# --- Peers · self-scoring percentile block (BELOW the auto table) -------------
# A DISTINCT second block: the subject's percentile within each cleaned peer
# column. Percentile is a 0-1 fraction; "Higher is" carries the polarity.
EXMPL_PEER_PCTL_HEADER = ["Metric", "Subject", "Peer median", "Percentile",
                          "Higher is", "Read"]
_WORSE = "worse (expensive/levered)"
_BETTER = "better"
EXMPL_PEER_PCTL = [
    # metric, subject, peer median, percentile (0-1), higher-is, read
    ("P/E (TTM)", 34.0, 37.5, 0.42, _WORSE, "mid-pack"),
    ("EV/EBITDA", 22.0, 24.0, 0.40, _WORSE, "mid-pack"),
    ("P/S (TTM)", 9.1, 5.4, 0.85, _WORSE, "expensive vs peers"),
    ("Rev growth % (3Y CAGR)", 18.0, 12.0, 0.80, _BETTER, "strong vs peers"),
    ("Gross margin %", 71.0, 62.0, 0.86, _BETTER, "strong vs peers"),
    ("Net margin %", 6.0, -4.0, 0.90, _BETTER, "strong vs peers"),
    ("FCF yield %", 2.1, 1.2, 0.75, _BETTER, "strong vs peers"),
    ("Net debt / EBITDA", 0.4, 1.8, 0.45, _WORSE, "mid-pack"),
]

# --- Valuation · auto WACC (CAPM) + reverse-DCF block -------------------------
# `None` value = a section/band title (col A only). The bare "WACC" row is the
# one the reader captures; the band title "WACC (auto — CAPM)" and the
# "WACC −1%"/"WACC +1%" sensitivity rows are decoys it must NOT match. Labels
# carry a Unicode minus (U+2212) exactly as the add-in emits them.
MINUS = "−"
EXMPL_VAL_WACC_DCF = [
    ("WACC (auto — CAPM)", None),
    ("Risk-free rate (10Y)", 0.043),
    ("Equity risk premium", 0.05),
    ("Beta (5y)", 1.6),
    ("Cost of equity (Ke)", 0.123),
    ("WACC", 0.092),
    (None, None),
    ("Reverse DCF — growth priced into today's price", None),
    (f"Implied growth  (WACC {MINUS} FCF/EV)", 0.061),
    ("Historical FCF CAGR (4y)", 0.045),
    ("Revenue CAGR (4y)", 0.12),
    (f"Priced-for {MINUS} delivered (FCF)", 0.016),
    ("Read", "market prices in MORE growth than history delivered"),
    (None, None),
    ("Implied growth sensitivity (WACC ±1%)", None),
    (f"WACC {MINUS}1%", 0.051),
    ("WACC base", 0.061),
    ("WACC +1%", 0.071),
]

# --- News · recent headlines (Date/Headline/Source/URL) -----------------------
# `None` = a blank spill row; NA in the headline = an unresolved row. Both are
# skipped on read. Sources are news outlets (Reuters/Bloomberg/AP/CNBC) — never
# a data-terminal brand.
EXMPL_NEWS_HEADER = ["Date", "Headline", "Source", "URL"]
EXMPL_NEWS = [
    (_dt.date(2026, 7, 14), "Example Corp posts record quarterly revenue, tops estimates",
     "Reuters", "https://example.com/news/record-revenue"),
    (_dt.date(2026, 7, 12), "Regulators open probe into logistics-software pricing",
     "Bloomberg", "https://example.com/news/pricing-probe"),
    None,  # blank spill row — must be skipped
    (_dt.date(2026, 7, 10), "Example Corp unveils next-gen routing platform at expo",
     "Associated Press", "https://example.com/news/routing-platform"),
    (_dt.date(2026, 7, 9), NA, "CNBC", "https://example.com/news/na-headline"),  # NA headline — skipped
    (_dt.date(2026, 7, 8), "CEO outlines margin recovery and cost cuts in interview",
     "CNBC", "https://example.com/news/ceo-interview"),
]


def build_workbook() -> Workbook:
    header_font = Font(bold=True)
    band = PatternFill("solid", fgColor="10331F")
    band_font = Font(bold=True, color="F2F2EF")
    wb = Workbook()

    # ---- Snapshot
    ws = wb.active
    ws.title = "Snapshot"
    ws["A1"] = "DENNIS — data · latest snapshot (formulas pre-filled)"
    ws["A2"], ws["B2"] = "Ticker / RIC ->", "EXMPL"
    hdr = ["Section", "field_key", "Label", "Value (auto)", "Field mnemonic (verify)", "Priority"]
    for j, h in enumerate(hdr, 1):
        ws.cell(row=4, column=j, value=h).font = header_font
    for i, (section, key, label, value, mnem) in enumerate(EXMPL_SNAPSHOT):
        r = 5 + i
        ws.cell(row=r, column=1, value=section)
        ws.cell(row=r, column=2, value=key)
        ws.cell(row=r, column=3, value=label)
        ws.cell(row=r, column=4, value=value)
        ws.cell(row=r, column=5, value=mnem)
        ws.cell(row=r, column=6, value="Required" if key in
                ("company_name", "ticker", "as_of_date", "price", "market_cap", "shares_out")
                else "Recommended")

    # ---- History
    hs = wb.create_sheet("History")
    hs["A1"] = "DENNIS — financial history (6 periods) + trend (formulas pre-filled)"
    head = ["field_key", "Label", *EXMPL_PERIODS, "CAGR FY-4->FY-0 %", "Field mnemonic (verify)"]
    for j, h in enumerate(head, 1):
        hs.cell(row=4, column=j, value=h).font = header_font
    for i, (key, label, vals) in enumerate(EXMPL_HISTORY):
        r = 5 + i
        hs.cell(row=r, column=1, value=key)
        hs.cell(row=r, column=2, value=label)
        for j, v in enumerate(vals):
            hs.cell(row=r, column=3 + j, value=v)
        hs.cell(row=r, column=3 + len(EXMPL_PERIODS), value="")  # CAGR (computed)
        hs.cell(row=r, column=4 + len(EXMPL_PERIODS), value="computed")

    # ---- Dashboard
    ds = wb.create_sheet("Dashboard")
    ds["A1"] = "DENNIS — one-glance summary (feeds the numbers sheet)"
    ds["A2"] = "Auto from the other sheets. Flags are rough heuristics."
    for i, (label, value) in enumerate(EXMPL_DASHBOARD):
        r = 3 + i
        ds.cell(row=r, column=1, value=label)
        ds.cell(row=r, column=2, value=value)
        if label == "Quick flags":
            ds.cell(row=r, column=1).fill = band
            ds.cell(row=r, column=1).font = band_font

    # ---- Valuation
    vs = wb.create_sheet("Valuation")
    vs["A1"] = "DENNIS — valuation & bear/base/bull"
    vs["A3"] = "Inputs & links"
    for i, (label, value) in enumerate(EXMPL_VAL_INPUTS):
        vs.cell(row=4 + i, column=1, value=label)
        vs.cell(row=4 + i, column=2, value=value)
    vhdr = ["Scenario", "Metric", "(EPS)", "Exit multiple", "Implied price", "Upside %"]
    for j, h in enumerate(vhdr, 1):
        vs.cell(row=7, column=j, value=h).font = header_font
    for i, sc in enumerate(EXMPL_VAL_SCENARIOS):
        for j, v in enumerate(sc):
            vs.cell(row=8 + i, column=1 + j, value=v)
    # WACC (CAPM) + reverse-DCF block, one blank row below the scenarios
    r0 = 8 + len(EXMPL_VAL_SCENARIOS) + 1
    for i, (label, value) in enumerate(EXMPL_VAL_WACC_DCF):
        if label is not None:
            vs.cell(row=r0 + i, column=1, value=label)
        if value is not None:
            vs.cell(row=r0 + i, column=2, value=value)

    # ---- Peers
    ps = wb.create_sheet("Peers")
    ps["A1"] = "DENNIS — peers (auto via PEERS(); optional / long-form)"
    for j, h in enumerate(EXMPL_PEERS_HEADER, 1):
        ps.cell(row=4, column=j, value=h).font = header_font
    for i, peer in enumerate(EXMPL_PEERS):
        for j, v in enumerate(peer):
            ps.cell(row=5 + i, column=1 + j, value=v)
    # self-scoring percentile block — a DISTINCT block beneath the auto table,
    # separated by one blank row so _read_peers stops before it.
    pp_title = 5 + len(EXMPL_PEERS) + 1
    ps.cell(row=pp_title, column=1,
            value="DENNIS — self-scoring vs peers (percentile of THIS ticker in each peer column)")
    ps.cell(row=pp_title + 1, column=1,
            value="PERCENTRANK.INC of the subject within the cleaned peer columns above. "
                  "'Read' bakes in direction so a glance tells you cheap/expensive and strong/weak.")
    pp_hdr = pp_title + 2
    for j, h in enumerate(EXMPL_PEER_PCTL_HEADER, 1):
        ps.cell(row=pp_hdr, column=j, value=h).font = header_font
    for i, mrow in enumerate(EXMPL_PEER_PCTL):
        for j, v in enumerate(mrow):
            ps.cell(row=pp_hdr + 1 + i, column=1 + j, value=v)
    # a footnote after a blank row — the reader must stop at the blank metric
    foot = pp_hdr + 1 + len(EXMPL_PEER_PCTL) + 1
    ps.cell(row=foot, column=1,
            value="Peer 3Y growth from revenue with date params; blank on refresh is harmless.")

    # ---- News
    ns = wb.create_sheet("News")
    ns["A1"] = "DENNIS — recent news & key developments (auto)"
    ns["A2"] = ("Latest headlines (Date / Headline / Source / URL), spilled by the add-in. "
                "A news outlet as Source is fine — never a data-terminal brand.")
    for j, h in enumerate(EXMPL_NEWS_HEADER, 1):
        ns.cell(row=4, column=j, value=h).font = header_font
    for i, item in enumerate(EXMPL_NEWS):
        if item is None:
            continue  # leave a genuinely blank spill row (exercises the skip)
        for j, v in enumerate(item):
            ns.cell(row=5 + i, column=1 + j, value=v)

    return wb


def gen_alignment_fixture(path: Path) -> None:
    """ElevenLabs with-timestamps style character alignment for a sample text."""
    text = "The market pays sixty times sales."
    words = text.split(" ")
    chars: list[str] = []
    starts: list[float] = []
    ends: list[float] = []
    t = 0.25  # small lead-in silence
    per_char = 0.045
    for i, w in enumerate(words):
        for ch in w:
            chars.append(ch)
            starts.append(round(t, 3))
            t += per_char
            ends.append(round(t, 3))
        if i < len(words) - 1:
            chars.append(" ")
            starts.append(round(t, 3))
            t += 0.06  # pause between words
            ends.append(round(t, 3))
    payload = {
        "text": text,
        "alignment": {
            "characters": chars,
            "character_start_times_seconds": starts,
            "character_end_times_seconds": ends,
        },
    }
    path.write_text(json.dumps(payload, indent=2))


def main() -> None:
    cdir = ROOT / "fixtures" / "company_data"
    adir = ROOT / "fixtures" / "tts"
    pdir = ROOT / "fixtures" / "prices"
    for d in (cdir, adir, pdir):
        d.mkdir(parents=True, exist_ok=True)

    build_workbook().save(cdir / "dennis_data.xlsx")

    csv_lines = ["field_key,value"]
    for _section, key, _label, value, _mnem in EXMPL_SNAPSHOT:
        csv_lines.append(f"{key},{value}")
    (cdir / "dennis_data.csv").write_text("\n".join(csv_lines) + "\n")

    gen_alignment_fixture(adir / "alignment_sample.json")

    from pipeline.prices import synthetic_series
    (pdir / "EXMPL.json").write_text(synthetic_series("EXMPL", 120).to_json())
    print("fixtures written")


if __name__ == "__main__":
    main()
