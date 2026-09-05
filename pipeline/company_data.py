"""Company-data export reader + filing-screenshot prep (§3).

The operator has an Excel add-in (Capital IQ, with Refinitiv/LSEG mnemonics
alongside), not API access, so the data contract is the shipped template — the
PRIVATE data source. The add-in resolves the formulas in Excel — since P3.1b
the bot can drive that itself; see `excel_refresh.py` — and the file this
reads carries CACHED VALUES, so it opens with openpyxl data_only=True
(values, never formula strings). Sheets are read strictly BY NAME, so the
add-in's hidden helper sheets (`_CIQHiddenCacheSheet`, `_RICMap`, a GUID-named
cache) are ignored:

  * `Snapshot`  — point-in-time: a `field_key` column + a value column,
    grouped by Section. The value column's title has changed between template
    revisions (`Value (auto)` → `Value (Capital IQ) MM`), so it is located by
    prefix — an exact match once read zero fields off a perfectly good
    workbook.
  * `History`   — 6 periods (FY-4 … FY-0, LTM) under the header row; the
    period columns are read DYNAMICALLY from that header, never hardcoded.
  * `Dashboard` — the one-glance summary + flags (this is exactly what the
    numbers sheet reads).
  * `Valuation` — bear/base/bull scenarios + inputs, plus the auto WACC
    (CAPM) and reverse-DCF block (long-form).
  * `Peers`     — the auto-pulled peer table, plus a self-scoring percentile
    block beneath it (two distinct blocks; long-form). Its metric columns are
    matched BY HEADER TEXT too: the table has grown revenue columns and a
    computed 3Y-CAGR since v3, and fixed offsets mislabelled everything to
    their right.
  * `News`      — recent headlines (Date/Headline/Source/URL), spilled by the
    add-in (optional; a news outlet as Source is fine — never a data-terminal
    brand).

A CSV export (`field_key,value`) is accepted for the snapshot only. Missing
required identity/size fields BLOCK the run; other gaps warn.

NOTHING here ever puts the vendor's name on screen: uploaded raw
screenshots are normalized with a generic "FROM THE 10-K" label.
"""

from __future__ import annotations

import csv
import datetime as _dt
import logging
from dataclasses import dataclass
from dataclasses import field as dc_field
from pathlib import Path

from openpyxl import load_workbook
from PIL import Image, ImageDraw, ImageFont

from config import Settings
from pipeline.models import (
    ALL_DATA_FIELDS,
    DATA_REQUIRED,
    HISTORY_FIELDS,
    CompanyData,
    _STRING_FIELDS,
)

log = logging.getLogger(__name__)

_NA_VALUES = {"", "#N/A", "#N/A N/A", "N/A", "NA", "#VALUE!", "#REF!", "#NAME?",
              "NULL", "-", "#DIV/0!"}
# the add-in leaves this literal in cells whose mnemonic didn't resolve
_FORMULA_ERR = "the formula must contain at least one field or function."

# --------------------------------------------------------------------------
# Cells the add-in did not resolve.
# --------------------------------------------------------------------------
# These are NOT the same thing as an empty cell, and the difference decides
# what the operator has to do about it. An empty cell means the vendor has no
# figure for this company — a real gap, and the run carries on with a warning.
# One of these means the refresh did not finish or the terminal was not signed
# in: the number exists, we just did not get it, and re-exporting is the fix.
#
# Left untreated they are considerably worse than a gap, because a text field
# accepts them. `company_name` reading `#CIQINACTIVE` is not missing, so it
# passes the required-field check and reaches the script — a video titled
# `#CIQINACTIVE`, built on numbers nobody looked at.
UNRESOLVED_MARKERS = (
    "#ciqinactive",          # CIQ: the add-in is loaded but not entitled/active
    "not signed in",         # CIQ/LSEG: no session
    "#name?",                # the add-in is not loaded at all
    "requesting data",       # still in flight when the sheet was saved
    "retrieving",
    "#getting_data",
    "#n/a n/a",              # LSEG's "no answer" shape
    "#n/a requesting data",
    "#value!",
    "#ref!",
    "#num!",
    "#null!",
    "#div/0!",
)
# `#N/A` alone is deliberately absent: on a values-only export it is the
# ordinary way a mnemonic says "no figure for this company", and treating it
# as a failure would reject perfectly good workbooks for thinly-covered
# small-caps. It is still NA (so it coerces to None), just not an alarm.


def unresolved_marker(raw) -> str | None:
    """The add-in failure this cell carries, or None.

    Returns the marker as written so the operator sees the thing that is
    actually in their sheet.
    """
    if raw is None or isinstance(raw, (int, float, _dt.date, _dt.datetime)):
        return None
    text = str(raw).strip()
    low = text.lower()
    for marker in UNRESOLVED_MARKERS:
        if low.startswith(marker) or marker in low:
            return text[:60]
    if low == _FORMULA_ERR:
        return text[:60]
    return None

# accepted upload names (the bot saves uploads under the first)
EXPORT_NAMES = ("dennis_data.xlsx", "data.xlsx", "dennis_data.csv")

# sheets read by name (anything else — Instructions, hidden helpers — ignored)
SNAPSHOT_SHEET = "Snapshot"
HISTORY_SHEET = "History"
DASHBOARD_SHEET = "Dashboard"
VALUATION_SHEET = "Valuation"
PEERS_SHEET = "Peers"
NEWS_SHEET = "News"


class CompanyDataError(Exception):
    pass


def _is_na(text: str) -> bool:
    t = text.strip()
    return t in _NA_VALUES or t.lower() == _FORMULA_ERR


def _num(raw) -> float | None:
    """NA-aware numeric coerce (values, never formulas)."""
    if raw is None:
        return None
    if isinstance(raw, (int, float)):
        return float(raw)
    text = str(raw).strip()
    if _is_na(text):
        return None
    try:
        return float(text.replace(",", "").replace("%", "").replace("$", ""))
    except ValueError:
        return None


def _coerce(field: str, raw) -> str | float | None:
    if raw is None:
        return None
    if isinstance(raw, (_dt.datetime, _dt.date)):
        # as_of_date etc. arrive as real dates — keep a clean ISO day
        return raw.date().isoformat() if isinstance(raw, _dt.datetime) else raw.isoformat()
    # An unresolved add-in cell is missing data, not data. Without this a text
    # field happily carries `#CIQINACTIVE` all the way onto the screen.
    marker = unresolved_marker(raw)
    if marker is not None:
        log.warning("company data field %s did not resolve (%s) — treating as "
                    "missing", field, marker)
        return None
    text = str(raw).strip()
    if _is_na(text):
        return None
    if field in _STRING_FIELDS:
        return text
    val = _num(raw)
    if val is None:
        log.warning("company data field %s: cannot coerce %r to number", field, raw)
    return val


def _text_or_none(raw) -> str | None:
    """NA-aware free-text coerce; a real date collapses to an ISO day string.
    Used by the News / peer-percentile readers, whose cells are free text (or
    spilled dates) rather than typed field_keys."""
    if raw is None:
        return None
    if isinstance(raw, _dt.datetime):
        return raw.date().isoformat()
    if isinstance(raw, _dt.date):
        return raw.isoformat()
    text = str(raw).strip()
    return None if _is_na(text) else text


def find_export(workspace: Path) -> Path | None:
    for name in EXPORT_NAMES:
        p = workspace / name
        if p.exists():
            return p
    return None


# --------------------------------------------------------------------------
# Upload validation — the primary data route.
# --------------------------------------------------------------------------
# The numbers arrive as a workbook the operator drops into the workspace:
# Excel is refreshed outside the bot and the result pasted into a clean sheet
# as VALUES. That makes this the front door, and a front door needs to say
# what is wrong in terms of the thing the operator can go and fix — not
# "missing required fields (company_name, ticker, …)", which is the same
# message whether they uploaded a stale file, the wrong company, or a
# workbook whose add-in never signed in.
#
# Five failures are worth naming separately, because each has a different fix:
#
#   missing sheet   -> re-export, the template lost a tab
#   formulas only   -> paste as values; the add-in's formulas do not travel
#   unresolved      -> sign the terminal in and refresh, then re-export
#   wrong ticker    -> you exported a different company
#   stale           -> refresh it; these numbers are from last month

MAX_REPORTED_FIELDS = 8


@dataclass(frozen=True)
class ExportProblem:
    """One specific, actionable thing wrong with an uploaded workbook."""

    kind: str            # missing_sheet | formulas_only | unresolved |
                         # wrong_ticker | stale | unreadable | no_fields
    message: str         # operator-facing; names the fix
    blocking: bool = True
    fields: tuple[str, ...] = ()


@dataclass
class ExportCheck:
    """What an uploaded workbook turned out to be."""

    path: Path
    ticker: str = ""
    as_of: str = ""
    sheets: tuple[str, ...] = ()
    problems: list[ExportProblem] = dc_field(default_factory=list)

    @property
    def ok(self) -> bool:
        return not self.blocking

    @property
    def blocking(self) -> list[ExportProblem]:
        return [p for p in self.problems if p.blocking]

    @property
    def warnings(self) -> list[ExportProblem]:
        return [p for p in self.problems if not p.blocking]

    def render(self) -> str:
        """The whole verdict, worst first, as chat text."""
        lines = [f"⛔ {p.message}" for p in self.blocking]
        lines += [f"⚠️ {p.message}" for p in self.warnings]
        return "\n".join(lines)


def _formula_cells(src: Path, sheet: str) -> int:
    """How many cells on `sheet` still hold a formula.

    A values-only paste has none. The reader loads with `data_only=True`, so
    a workbook that still carries live add-in formulas reads as *empty* unless
    Excel happened to cache results before saving — the difference between
    "no data" and "the data is there but not in a form we can read", which are
    very different things to be told.
    """
    try:
        wb = load_workbook(src, data_only=False, read_only=True)
    except Exception:  # noqa: BLE001 - unreadable is reported by the caller
        return 0
    try:
        if sheet not in wb.sheetnames:
            return 0
        return sum(
            1
            for row in wb[sheet].iter_rows()
            for c in row
            if isinstance(c.value, str) and c.value.startswith("=")
        )
    finally:
        wb.close()


def check_export(
    src: Path,
    *,
    expect_ticker: str = "",
    max_age_days: int | None = None,
    today: _dt.date | None = None,
) -> ExportCheck:
    """Validate an uploaded workbook and say precisely what is wrong with it.

    Never raises: an unreadable file is itself one of the answers.
    """
    src = Path(src)
    check = ExportCheck(path=src)

    if src.suffix.lower() == ".csv":
        # The CSV route carries the snapshot only, so there are no sheets to
        # check and no formulas to have failed to paste. Everything that is
        # about the *contents* still applies though: a CSV for the wrong
        # company overwrites the right one just as thoroughly.
        try:
            pairs = _read_csv_pairs(src)
        except OSError as e:
            check.problems.append(ExportProblem(
                kind="unreadable",
                message=f"{src.name} could not be read ({e})."))
            return check
        if not pairs:
            check.problems.append(ExportProblem(
                kind="no_fields",
                message=(f"{src.name} has no `field_key,value` rows. A CSV "
                         f"export carries the snapshot only, one field per "
                         f"line.")))
            return check
        _check_contents(check, pairs, expect_ticker, max_age_days, today)
        return check

    try:
        wb = load_workbook(src, data_only=True)
    except Exception as e:  # noqa: BLE001 - openpyxl raises a zoo of these
        check.problems.append(ExportProblem(
            kind="unreadable",
            message=(f"{src.name} could not be opened as a workbook ({e}). "
                     f"Re-save it as .xlsx and upload it again.")))
        return check

    try:
        check.sheets = tuple(wb.sheetnames)
        if SNAPSHOT_SHEET not in wb.sheetnames:
            visible = ", ".join(n for n in wb.sheetnames if not n.startswith("_")) or "none"
            check.problems.append(ExportProblem(
                kind="missing_sheet",
                message=(f"no '{SNAPSHOT_SHEET}' sheet in {src.name} — that is "
                         f"where every field is read from. Sheets present: "
                         f"{visible}. Re-export from the data template."),
                fields=(SNAPSHOT_SHEET,)))
            return check

        pairs = _read_snapshot(wb[SNAPSHOT_SHEET])
        # A formula cell with no cached result reads as None under
        # `data_only=True`, so an unpasted workbook yields keys with nothing
        # behind them — not an empty dict. Both are "nothing was read".
        has_values = any(v is not None and str(v).strip() for v in pairs.values())

        if not has_values:
            formulas = _formula_cells(src, SNAPSHOT_SHEET)
            if formulas:
                check.problems.append(ExportProblem(
                    kind="formulas_only",
                    message=(f"{src.name} still contains {formulas} add-in "
                             f"formula(s) and no saved values. The formulas do "
                             f"not travel — they only resolve on a machine with "
                             f"the add-in signed in. Copy the sheet and Paste "
                             f"Special ▸ Values into a clean workbook, then "
                             f"upload that.")))
            elif not pairs:
                check.problems.append(ExportProblem(
                    kind="no_fields",
                    message=(f"the '{SNAPSHOT_SHEET}' sheet in {src.name} has no "
                             f"'field_key' column, so nothing could be read. "
                             f"Re-export from the data template.")))
            else:
                check.problems.append(ExportProblem(
                    kind="no_fields",
                    message=(f"every field in {src.name} is empty. If the "
                             f"workbook looks right on screen, it was probably "
                             f"saved before the add-in finished — refresh it, "
                             f"wait for the cells to settle, and re-export.")))
            return check

        _check_contents(check, pairs, expect_ticker, max_age_days, today)
        return check
    finally:
        wb.close()


def _read_csv_pairs(src: Path) -> dict[str, object]:
    """`field_key,value` rows from a CSV export. utf-8-sig: Excel writes a BOM."""
    pairs: dict[str, object] = {}
    with open(src, newline="", encoding="utf-8-sig") as f:
        for row in csv.reader(f):
            if len(row) >= 2 and row[0].strip() and row[0].strip() not in ("field", "field_key"):
                pairs[row[0].strip()] = row[1]
    return pairs


# Exchange codes a RIC may carry. A curated list rather than "any short
# alphabetic tail", because `.A` and `.B` are share classes: BRK.A and BRK.B
# are different securities and must not collapse onto each other.
RIC_SUFFIXES = frozenset({
    "O", "OQ", "N", "P", "K", "Q", "PK", "V",          # US
    "L", "PA", "DE", "F", "MI", "MC", "AS", "BR",      # Europe
    "SW", "ST", "HE", "CO", "OL", "VI", "LS", "IR",
    "TO", "TSX", "MX", "SA", "BA", "SN",               # Americas
    "AX", "NZ", "T", "HK", "SS", "SZ", "KS", "KQ",     # Asia-Pacific
    "TW", "TWO", "SI", "BO", "NS", "JK", "BK", "KL",
    "SR", "TA", "JO", "CA",                            # MEA
})


def ticker_root(value: str) -> str:
    """A ticker with its exchange/vendor suffix removed.

    `SNDK.O`, `SNDK.OQ`, `VOD.L` and `SNDK` all name the same company. The
    terminal writes a RIC, the operator opens a workspace with a plain ticker,
    and an exact string comparison rejected a workbook that was for exactly
    the right company — which is how a valid export got refused at upload.

    Only a suffix in :data:`RIC_SUFFIXES` is stripped. `BRK.B` keeps its tail:
    a share class is a different security, and quietly treating it as the same
    one would swap the wrong company's numbers into a video.
    """
    head, _, tail = value.strip().upper().partition(".")
    return head if tail in RIC_SUFFIXES else value.strip().upper()


def tickers_match(a: str, b: str) -> bool:
    """True when two ticker spellings name the same company."""
    a, b = a.strip().upper(), b.strip().upper()
    return bool(a) and bool(b) and (a == b or ticker_root(a) == ticker_root(b))


def _check_contents(check: ExportCheck, pairs: dict, expect_ticker: str,
                    max_age_days: int | None, today: _dt.date | None) -> None:
    """The checks that are about what the export SAYS, not how it is shaped.

    Shared by the workbook and CSV routes: a CSV for the wrong company
    overwrites the right one exactly as thoroughly as a workbook does.
    """
    name = check.path.name
    check.ticker = str(_coerce("ticker", pairs.get("ticker")) or "")
    check.as_of = str(_coerce("as_of_date", pairs.get("as_of_date")) or "")

    # --- unresolved cells ---------------------------------------------
    unresolved = {k: m for k in pairs
                  if (m := unresolved_marker(pairs.get(k))) is not None}
    required_unresolved = [k for k in DATA_REQUIRED if k in unresolved]
    other_unresolved = [k for k in sorted(unresolved) if k not in DATA_REQUIRED]

    if required_unresolved:
        shown = ", ".join(f"{k} ({unresolved[k]})"
                          for k in required_unresolved[:MAX_REPORTED_FIELDS])
        check.problems.append(ExportProblem(
            kind="unresolved",
            message=(f"required field(s) never resolved in {name}: "
                     f"{shown}. That is the add-in reporting a problem, not "
                     f"a company with no data — check the terminal is "
                     f"signed in, refresh, and re-export."),
            fields=tuple(required_unresolved)))
    if other_unresolved:
        shown = ", ".join(other_unresolved[:MAX_REPORTED_FIELDS])
        more = (f" …and {len(other_unresolved) - MAX_REPORTED_FIELDS} more"
                if len(other_unresolved) > MAX_REPORTED_FIELDS else "")
        check.problems.append(ExportProblem(
            kind="unresolved", blocking=False,
            message=(f"{len(other_unresolved)} optional field(s) did not "
                     f"resolve: {shown}{more}. They are treated as missing."),
            fields=tuple(other_unresolved)))

    # --- wrong company -------------------------------------------------
    want = expect_ticker.strip().upper()
    got = check.ticker.strip().upper()
    if want and got and not tickers_match(got, want):
        check.problems.append(ExportProblem(
            kind="wrong_ticker",
            message=(f"{name} is {got}, but this workspace is {want}. "
                     f"Nothing was overwritten. Export {want} and upload "
                     f"that, or start a workspace for {got} with "
                     f"/short {got} or /long {got}."),
            fields=(got,)))
    elif want and not got and "ticker" not in unresolved:
        check.problems.append(ExportProblem(
            kind="wrong_ticker", blocking=False,
            message=(f"{name} carries no ticker, so it cannot be "
                     f"confirmed as {want}.")))

    # --- staleness ------------------------------------------------------
    # Keyed off the export's own as-of date. This is the authority now: the
    # numbers are refreshed outside the bot, so nothing else here knows when
    # they were pulled, and a file's mtime only records the last save.
    if max_age_days is None:
        return
    age = _as_of_age_days(check.as_of, today)
    if age is None and not check.as_of:
        check.problems.append(ExportProblem(
            kind="stale", blocking=False,
            message=f"{name} carries no as-of date, so its age cannot be checked."))
    elif age is None:
        check.problems.append(ExportProblem(
            kind="stale", blocking=False,
            message=f"could not read the as-of date {check.as_of!r} in {name}."))
    elif age > max_age_days:
        check.problems.append(ExportProblem(
            kind="stale", blocking=False,
            message=(f"{name} is {age} days old (as of {check.as_of}; limit "
                     f"{max_age_days}). Refresh it before rendering — a video "
                     f"states these as current numbers.")))


def _as_of_age_days(as_of: str, today: _dt.date | None = None) -> int | None:
    """Age in days of an as-of date string, or None if it cannot be read."""
    if not as_of:
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"):
        try:
            parsed = _dt.datetime.strptime(as_of.strip()[:10], fmt).date()
        except ValueError:
            continue
        return ((today or _dt.date.today()) - parsed).days
    return None


# --------------------------------------------------------------------------
# v3 sheet readers — locate headers by text, never by fixed cell position.
# --------------------------------------------------------------------------

def _rows(ws) -> list[tuple]:
    return list(ws.iter_rows(values_only=True))


def _cell_text(v) -> str:
    return "" if v is None else str(v).strip()


def _header_row(rows: list[tuple], marker: str) -> int | None:
    """Index of the first row containing a cell equal to `marker`."""
    m = marker.strip().lower()
    for i, row in enumerate(rows):
        if any(_cell_text(c).lower() == m for c in row):
            return i
    return None


def _col_of(header: tuple, *names: str) -> int | None:
    """Column index whose header matches one of `names`.

    Exact match first, then a prefix match. The prefix pass matters: the value
    column is titled `Value (auto)` in one template revision and
    `Value (Capital IQ) MM` in the next, and an exact-only match silently read
    ZERO snapshot fields off the newer one — a workbook that looked present
    and parsed to nothing.
    """
    wanted = [n.strip().lower() for n in names]
    cells = [_cell_text(c).lower() for c in header]
    for j, text in enumerate(cells):
        if text in wanted:
            return j
    for j, text in enumerate(cells):
        if text and any(text.startswith(w) for w in wanted):
            return j
    return None


def _read_snapshot(ws) -> dict[str, object]:
    rows = _rows(ws)
    hr = _header_row(rows, "field_key")
    if hr is None:
        return {}
    header = rows[hr]
    key_c = _col_of(header, "field_key")
    val_c = _col_of(header, "value (auto)", "value")
    if key_c is None or val_c is None:
        return {}
    pairs: dict[str, object] = {}
    for row in rows[hr + 1:]:
        if key_c >= len(row):
            continue
        key = _cell_text(row[key_c])
        if not key or key == "field_key":
            continue
        pairs[key] = row[val_c] if val_c < len(row) else None
    return pairs


def _read_history(ws) -> tuple[list[str], dict[str, list[float | None]]]:
    """Header row carries `field_key`, `Label`, the period labels, then
    `CAGR …` and the mnemonic column. The period columns are everything
    between `Label` and the CAGR/mnemonic tail — read dynamically."""
    rows = _rows(ws)
    hr = _header_row(rows, "field_key")
    if hr is None:
        return [], {}
    header = rows[hr]
    key_c = _col_of(header, "field_key")
    label_c = _col_of(header, "label")
    start = (label_c if label_c is not None else key_c) + 1
    period_cols: list[int] = []
    periods: list[str] = []
    for j in range(start, len(header)):
        txt = _cell_text(header[j])
        low = txt.lower()
        if not txt:
            continue
        if low.startswith("cagr") or "mnemonic" in low:
            break  # the trailing computed/verify columns
        period_cols.append(j)
        periods.append(txt)
    if not period_cols:
        return [], {}
    history: dict[str, list[float | None]] = {}
    for row in rows[hr + 1:]:
        if key_c >= len(row):
            continue
        key = _cell_text(row[key_c])
        if not key or key == "field_key":
            continue
        history[key] = [_num(row[j] if j < len(row) else None) for j in period_cols]
    unknown = [k for k in history if k not in HISTORY_FIELDS]
    for k in unknown:
        log.warning("history sheet has unknown field %r — ignored", k)
        history.pop(k)
    return periods, history


def _read_dashboard(ws) -> dict[str, object]:
    """`label | value` pairs (the one-glance summary + flags). Skips the
    title/subtitle banner rows; keeps `yes/no` flag strings verbatim."""
    out: dict[str, object] = {}
    for row in _rows(ws):
        if not row:
            continue
        label = _cell_text(row[0])
        if not label:
            continue
        low = label.lower()
        if low.startswith("dennis") or low.startswith("auto from"):
            continue
        raw = row[1] if len(row) > 1 else None
        if _cell_text(raw) == "" or label.lower() == "quick flags":
            out[label] = None if label.lower() != "quick flags" else ""
            continue
        text = _cell_text(raw)
        if _is_na(text):
            out[label] = None
        elif text.lower() in ("yes", "no"):
            out[label] = text.lower()
        else:
            num = _num(raw)
            out[label] = num if num is not None else text
    return out


def _read_news(ws) -> list[dict]:
    """Recent headlines. The header row carries `Date | Headline | Source |
    URL` (anchor on the row with a cell equal to `Headline`); data rows spill
    beneath, possibly blank/NA. Skip any row whose headline is empty/NA;
    dates coerce to ISO day strings. Source is a news outlet, never a
    data-terminal brand."""
    rows = _rows(ws)
    hr = _header_row(rows, "headline")
    if hr is None:
        return []
    header = rows[hr]
    date_c = _col_of(header, "date")
    head_c = _col_of(header, "headline")
    src_c = _col_of(header, "source")
    url_c = _col_of(header, "url")
    if head_c is None:
        return []

    def _cell(row, c):
        return row[c] if c is not None and c < len(row) else None

    news: list[dict] = []
    for row in rows[hr + 1:]:
        headline = _text_or_none(_cell(row, head_c))
        if not headline:
            continue  # blank/NA spill row — skip (do NOT stop; gaps happen)
        news.append({
            "date": _text_or_none(_cell(row, date_c)),
            "headline": headline,
            "source": _text_or_none(_cell(row, src_c)),
            "url": _text_or_none(_cell(row, url_c)),
        })
    return news


def _read_valuation(ws) -> dict:
    rows = _rows(ws)
    val: dict = {}

    def _find_value(label: str):
        for row in rows:
            if row and _cell_text(row[0]).lower() == label:
                return _num(row[1]) if len(row) > 1 else None
        return None

    def _find_value_prefix(prefix: str):
        """First row whose col-A label STARTS WITH `prefix` (lowercased). The
        add-in emits these labels with a Unicode minus (U+2212) and en-dashes,
        so we anchor on the clean leading text, never on hardcoded punctuation."""
        for row in rows:
            if row and _cell_text(row[0]).lower().startswith(prefix):
                return _num(row[1]) if len(row) > 1 else None
        return None

    def _find_text(label: str):
        for row in rows:
            if row and _cell_text(row[0]).lower() == label:
                return _text_or_none(row[1]) if len(row) > 1 else None
        return None

    val["current_price"] = _find_value("current price")
    val["ltm_eps"] = _find_value("ltm eps")
    val["ltm_fcf_ps"] = _find_value("ltm fcf / share")

    # WACC + reverse-DCF block (fully auto). Match on exact / leading text so
    # the bare `WACC` row is captured but NOT the "WACC (auto — CAPM)" band
    # title nor the "WACC −1%" / "WACC +1%" sensitivity rows. `implied_growth`
    # takes the reverse-DCF row (first "Implied growth …"), not the later
    # "Implied growth sensitivity …" title.
    val["wacc"] = _find_value("wacc")
    val["implied_growth"] = _find_value_prefix("implied growth")
    val["hist_fcf_cagr"] = _find_value("historical fcf cagr (4y)")
    val["rev_cagr"] = _find_value("revenue cagr (4y)")
    val["priced_vs_delivered"] = _find_value_prefix("priced-for")
    val["reverse_dcf_read"] = _find_text("read")

    hr = _header_row(rows, "scenario")
    scenarios: list[dict] = []
    if hr is not None:
        for row in rows[hr + 1:]:
            name = _cell_text(row[0]) if row else ""
            if name.lower() not in ("bear", "base", "bull"):
                continue
            scenarios.append({
                "scenario": name,
                "metric": _cell_text(row[1]) or _cell_text(row[2]) if len(row) > 2 else "",
                "exit_multiple": _num(row[3]) if len(row) > 3 else None,
                "implied_price": _num(row[4]) if len(row) > 4 else None,
                "upside_pct": _num(row[5]) if len(row) > 5 else None,
            })
    val["scenarios"] = scenarios
    return val


# Peer-table fields -> the header texts that carry them, in sheet order.
# Located BY HEADER TEXT like every other reader here, never by position: the
# table grew `Rev LTM (now)` / `Rev LTM (-3Y)` feed columns and moved revenue
# growth to a COMPUTED 3Y CAGR at the far right, which shifted everything from
# the margins rightward by one column. Read off fixed offsets that silently
# mislabelled five fields — margins landing in `rev_growth`, and a raw revenue
# figure landing in `net_debt_ebitda` — numbers that were plausible enough to
# reach a script. Aliases cover older revisions (a raw `Rev Growth %` column,
# spelled-out margin titles); `_col_of` matches exact first and only then by
# prefix, so the `cln …` helper columns to the right are never picked up.
_PEER_COLS: dict[str, tuple[str, ...]] = {
    "price": ("price",),
    "market_cap": ("market cap",),
    "pe": ("p/e",),
    "ev_ebitda": ("ev/ebitda",),
    "ps": ("p/s",),
    "gross_margin": ("gross mgn %", "gross margin %"),
    "net_margin": ("net mgn %", "net margin %"),
    "fcf_yield": ("fcf yield %",),
    "net_debt_ebitda": ("netdebt/ebitda", "net debt/ebitda", "net debt / ebitda"),
    # the COMPUTED 3Y CAGR — never the raw `Rev LTM …` columns that feed it
    "rev_growth": ("rev growth % (3y cagr)", "rev growth %", "revenue growth %"),
}


def _read_peers(ws) -> list[dict]:
    """The auto peer table — the FIRST block on the Peers sheet (the
    self-scoring percentile block beneath it is `_read_peer_percentiles`).
    Anchor on the header row carrying `Peer (auto)` and take every metric
    column from that header by text; a column the template doesn't carry
    reads None rather than dragging its neighbour in."""
    rows = _rows(ws)
    hr = _header_row(rows, "peer (auto)")
    if hr is None:
        return []
    header = rows[hr]
    name_c = _col_of(header, "peer (auto)", "peer")
    if name_c is None:
        name_c = 0
    cols = {field: _col_of(header, *names) for field, names in _PEER_COLS.items()}
    absent = [f for f, c in cols.items() if c is None]
    if absent:
        log.warning("peers sheet has no column for %s — left empty", ", ".join(absent))
    peers: list[dict] = []
    for row in rows[hr + 1:]:
        name = _cell_text(row[name_c]) if name_c < len(row) else ""
        if not name:
            # the PEERS() spill is contiguous; the first blank name ends the
            # auto table — stop here so we never wander into the self-scoring
            # percentile block that lives beneath it.
            break
        entry: dict = {"name": name}
        for field, c in cols.items():
            entry[field] = _num(row[c]) if c is not None and c < len(row) else None
        peers.append(entry)
    return peers


def _percentile_direction(raw) -> str | None:
    """Normalize the "Higher is" column to `better` / `worse`. The cell reads
    e.g. "better" or "worse (expensive/levered)" — we keep just the polarity."""
    text = _text_or_none(raw)
    if text is None:
        return None
    low = text.lower()
    if low.startswith("better"):
        return "better"
    if low.startswith("worse"):
        return "worse"
    return None


def _read_peer_percentiles(ws) -> list[dict]:
    """The SECOND block on the Peers sheet — the subject's self-score vs its
    peers — which lives beneath the auto table (`_read_peers` handles that).
    Anchor on the header row whose col-A cell equals `Metric`
    (Metric | Subject | Peer median | Percentile | Higher is | Read | Peer low
    | Peer high | t (subject) | t (median)); metric rows follow until the first
    blank metric. `percentile` is a 0–1 fraction; `direction` is the polarity
    of the "Higher is" column; `read` is the plain-text verdict ("expensive vs
    peers" / "strong vs peers" / …).

    **`t` IS NOT `percentile`, AND SUBSTITUTING ONE FOR THE OTHER IS THE
    MISTAKE THIS BLOCK IS SHAPED TO PREVENT.** `tables/multiples-strip`'s rail
    is a VALUE axis: `t` is where the subject's number falls between the peer
    low and the peer high, which is what `Peers!I` and `Peers!J` compute.
    `percentile` (column D) is RANK-based, and feeding it to the rail computes
    `t_median` to 0.5 on every row — every median tick dead centre, which is
    both wrong and the death of the one comparison the plate exists to make.

    The ends are p10 and p90 rather than min and max on purpose: one peer on a
    900x P/E would otherwise own the axis and squash every subject to t = 0.02.
    **So `t` can legitimately fall outside 0–1, and nothing here clamps it** —
    off the peer range is the most quotable row on the plate, and the renderer
    draws it as a dot on the end tick with a chevron past it.

    Column E ("higher is better/worse") is for the caption, never the rail: the
    rail is a position, and a position is not inverted by the metric's polarity.
    """
    rows = _rows(ws)
    hr = _header_row(rows, "metric")
    if hr is None:
        return []
    header = rows[hr]
    m_c = _col_of(header, "metric")
    subj_c = _col_of(header, "subject")
    med_c = _col_of(header, "peer median")
    pct_c = _col_of(header, "percentile")
    dir_c = _col_of(header, "higher is")
    read_c = _col_of(header, "read")
    low_c = _col_of(header, "peer low")
    high_c = _col_of(header, "peer high")
    t_c = _col_of(header, "t (subject)")
    tmed_c = _col_of(header, "t (median)")
    if m_c is None:
        return []

    def _cell(row, c):
        return row[c] if c is not None and c < len(row) else None

    out: list[dict] = []
    for row in rows[hr + 1:]:
        metric = _cell_text(row[m_c]) if m_c < len(row) else ""
        if not metric:
            break  # the first blank metric ends the block
        out.append({
            "metric": metric,
            "subject": _num(_cell(row, subj_c)),
            "median": _num(_cell(row, med_c)),
            "percentile": _num(_cell(row, pct_c)),
            "direction": _percentile_direction(_cell(row, dir_c)),
            "read": _text_or_none(_cell(row, read_c)),
            # The rail. Read as written, unclamped — see the docstring.
            "peer_low": _num(_cell(row, low_c)),
            "peer_high": _num(_cell(row, high_c)),
            "t": _num(_cell(row, t_c)),
            "t_median": _num(_cell(row, tmed_c)),
        })
    return out


def load_company_data(workspace: Path) -> CompanyData:
    """Load + type-coerce the v3 export (all sheets, by name). Raises
    CompanyDataError if absent or unreadable; missing-field policy lives on
    the model."""
    src = find_export(workspace)
    if src is None:
        raise CompanyDataError(
            "No dennis_data.xlsx / data.xlsx / dennis_data.csv in the workspace. "
            "Refresh the data template for this ticker and upload it."
        )

    pairs: dict[str, object] = {}
    history_years: list[str] = []
    history: dict[str, list[float | None]] = {}
    dashboard: dict[str, object] = {}
    valuation: dict = {}
    peers: list[dict] = []
    peer_percentiles: list[dict] = []
    news: list[dict] = []
    if src.suffix == ".xlsx":
        wb = load_workbook(src, data_only=True)
        names = set(wb.sheetnames)
        snap_ws = wb[SNAPSHOT_SHEET] if SNAPSHOT_SHEET in names else wb.worksheets[0]
        pairs = _read_snapshot(snap_ws)
        if HISTORY_SHEET in names:
            history_years, history = _read_history(wb[HISTORY_SHEET])
        else:
            log.warning("export has no History sheet — multi-year numbers unavailable")
        if DASHBOARD_SHEET in names:
            dashboard = _read_dashboard(wb[DASHBOARD_SHEET])
        if VALUATION_SHEET in names:
            valuation = _read_valuation(wb[VALUATION_SHEET])
        if PEERS_SHEET in names:
            # two distinct blocks on the one sheet — the raw auto table and
            # the self-scoring percentile block below it.
            peers = _read_peers(wb[PEERS_SHEET])
            peer_percentiles = _read_peer_percentiles(wb[PEERS_SHEET])
        if NEWS_SHEET in names:
            news = _read_news(wb[NEWS_SHEET])
        else:
            log.warning("export has no News sheet — recent-headlines flow unavailable")
        wb.close()
    else:
        pairs = _read_csv_pairs(src)
        log.warning("CSV export carries the snapshot only — no history/dashboard")

    unknown = [k for k in pairs if k not in ALL_DATA_FIELDS]
    for k in unknown:
        log.warning("export has unknown field %r — ignored", k)

    values = {field: _coerce(field, pairs.get(field)) for field in ALL_DATA_FIELDS}
    return CompanyData(values=values, history_years=history_years,
                       history=history, dashboard=dashboard,
                       valuation=valuation, peers=peers,
                       peer_percentiles=peer_percentiles, news=news,
                       source_file=str(src))


# ---------------------------------------------------------------------------
# Screenshot prep: raw data screenshots -> normalized full-screen flashes
# with a GENERIC source label — the vendor is never named on screen (§3).
# ---------------------------------------------------------------------------

FILING_LABEL = "FROM THE 10-K"


def prepare_screenshot(src: Path, dest: Path, settings: Settings) -> Path:
    """Full-frame designed filing card: the screenshot fitted sharp over a
    blurred, brand-tinted cover of itself (never a letterboxed black frame),
    a subtle border, and the generic '10-K' chip. Deterministic output."""
    from pipeline.rasters import cover_fill_frame, role

    W, H = settings.long_resolution
    margin = int(H * 0.05)
    canvas = cover_fill_frame(src, W, H, keep_min=1.1,   # always contain-on-fill
                              ground=role(settings, "ground"),
                              line=role(settings, "structure"))
    d = ImageDraw.Draw(canvas)

    # generic source chip — "the filing", never the vendor
    font = ImageFont.truetype(str(settings.fonts_dir / "SpaceMono-Bold.ttf"),
                              max(int(H * 0.026), 14))
    pad = int(H * 0.012)
    tw = d.textlength(FILING_LABEL, font=font)
    cx0, cy0 = margin // 2, margin // 2
    d.rounded_rectangle(
        [cx0, cy0, cx0 + tw + 2 * pad, cy0 + font.size + 2 * pad],
        radius=8, fill=(24, 28, 36), outline=(96, 106, 122), width=2,
    )
    d.text((cx0 + pad, cy0 + pad), FILING_LABEL, font=font, fill=(47, 213, 118))

    dest.parent.mkdir(parents=True, exist_ok=True)
    canvas.save(dest)
    return dest


def list_screenshots(workspace: Path) -> list[str]:
    """Raw screenshot files the operator dropped into the workspace (these
    are what [SHOW FILING: file] may reference)."""
    reserved = set(EXPORT_NAMES)
    return sorted(
        p.name for p in workspace.glob("*.png")
        if p.name not in reserved and not p.name.startswith("_")
    )
