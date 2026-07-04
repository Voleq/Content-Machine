"""Refinitiv export reader + screenshot prep (§5.1).

The operator has the LSEG Excel add-in, not API access, so the data
contract is `templates/refinitiv_audit_template.xlsx`: a fixed "Audit"
sheet where column A holds the stable field name and column B the value
(filled by Refinitiv formulas). We read strictly by field name — never by
guessing layout. A CSV export (`field,value` columns) is accepted too.

Missing identity/size/margins/cash fields BLOCK the run; other gaps warn.
"""

from __future__ import annotations

import csv
import logging
from pathlib import Path

from openpyxl import load_workbook
from PIL import Image, ImageDraw

from config import Settings
from pipeline.models import ALL_REFINITIV_FIELDS, RefinitivAudit, _STRING_FIELDS

log = logging.getLogger(__name__)

_NA_VALUES = {"", "#N/A", "#N/A N/A", "N/A", "NA", "#VALUE!", "#REF!", "#NAME?", "NULL", "-"}


class RefinitivError(Exception):
    pass


def _coerce(field: str, raw) -> str | float | None:
    if raw is None:
        return None
    text = str(raw).strip()
    if text in _NA_VALUES:
        return None
    if field in _STRING_FIELDS:
        return text
    try:
        return float(text.replace(",", "").replace("%", "").replace("$", ""))
    except ValueError:
        log.warning("refinitiv field %s: cannot coerce %r to number", field, raw)
        return None


def find_export(workspace: Path) -> Path | None:
    for name in ("data_refinitiv.xlsx", "data_refinitiv.csv"):
        p = workspace / name
        if p.exists():
            return p
    return None


def load_audit(workspace: Path) -> RefinitivAudit:
    """Load + type-coerce the export. Raises RefinitivError if absent or
    unreadable; missing-field policy is exposed on the returned model."""
    src = find_export(workspace)
    if src is None:
        raise RefinitivError(
            "No data_refinitiv.xlsx / data_refinitiv.csv in the workspace. "
            "Refresh the audit template for this ticker and upload it."
        )

    pairs: dict[str, object] = {}
    if src.suffix == ".xlsx":
        wb = load_workbook(src, data_only=True, read_only=True)
        ws = wb["Audit"] if "Audit" in wb.sheetnames else wb.worksheets[0]
        for row in ws.iter_rows(min_row=1, max_col=2, values_only=True):
            if not row or row[0] is None:
                continue
            field = str(row[0]).strip()
            if field and field != "field":
                pairs[field] = row[1] if len(row) > 1 else None
        wb.close()
    else:
        with open(src, newline="", encoding="utf-8-sig") as f:
            for row in csv.reader(f):
                if len(row) >= 2 and row[0].strip() and row[0].strip() != "field":
                    pairs[row[0].strip()] = row[1]

    unknown = [k for k in pairs if k not in ALL_REFINITIV_FIELDS]
    for k in unknown:
        log.warning("refinitiv export has unknown field %r — ignored", k)

    values = {
        field: _coerce(field, pairs.get(field))
        for field in ALL_REFINITIV_FIELDS
    }
    return RefinitivAudit(values=values, source_file=str(src))


# ---------------------------------------------------------------------------
# Screenshot prep: raw Refinitiv screenshots -> normalized full-screen
# "raw reality" flash overlays (§5.1, §7.2).
# ---------------------------------------------------------------------------


def prepare_screenshot(src: Path, dest: Path, settings: Settings) -> Path:
    """Fixed canvas, fitted image, subtle border — deterministic output."""
    W, H = settings.long_resolution
    margin = int(H * 0.05)
    img = Image.open(src).convert("RGB")
    img.thumbnail((W - 2 * margin, H - 2 * margin), Image.LANCZOS)

    canvas = Image.new("RGB", (W, H), (10, 13, 18))
    x = (W - img.width) // 2
    y = (H - img.height) // 2
    canvas.paste(img, (x, y))
    d = ImageDraw.Draw(canvas)
    d.rectangle([x - 3, y - 3, x + img.width + 2, y + img.height + 2],
                outline=(96, 106, 122), width=3)
    dest.parent.mkdir(parents=True, exist_ok=True)
    canvas.save(dest)
    return dest


def list_screenshots(workspace: Path) -> list[str]:
    """Raw screenshot files the operator dropped into the workspace (these
    are what [SHOW REFINITIV: file] may reference)."""
    reserved = {"data_refinitiv.xlsx", "data_refinitiv.csv"}
    return sorted(
        p.name for p in workspace.glob("*.png")
        if p.name not in reserved and not p.name.startswith("_")
    )
